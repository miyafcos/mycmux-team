from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
QA_DIR = REPO_ROOT / "docs" / "qa"
WORKBOOK = QA_DIR / "mycmux-feature-status-canonical.xlsx"

REQUIRED_SHEETS = {
    "Overview",
    "Feature Status",
    "Test Matrix",
    "Error Log",
    "Manual Retest Queue",
    "Safe Window Playbook",
    "Evidence Index",
    "Validation Log",
    "Test Run Log",
}

REQUIRED_QUEUE_BUCKETS = {
    "reload-required-live-fix-retest",
    "restart-or-duplicate-launch-window",
    "side-effect-gated",
    "interactive-live-ui-window",
    "manual-runtime-window",
    "backlog-confirmation",
}


def rows_by_header(sheet) -> list[dict[str, object]]:
    headers = [str(cell.value) for cell in sheet[1]]
    return [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]


def test_qa_tracker_is_single_canonical_workbook() -> None:
    workbooks = sorted(path.name for path in QA_DIR.glob("*.xlsx"))
    assert workbooks == ["mycmux-feature-status-canonical.xlsx"]


def test_qa_tracker_required_sheets_and_story_alignment() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    assert REQUIRED_SHEETS.issubset(set(wb.sheetnames))

    feature_rows = rows_by_header(wb["Feature Status"])
    matrix_rows = rows_by_header(wb["Test Matrix"])
    feature_ids = {str(row["Story ID"]) for row in feature_rows}
    matrix_ids = {str(row["Story ID"]) for row in matrix_rows}

    assert len(feature_ids) == 56
    assert feature_ids == matrix_ids
    assert all(str(row["User Story"]).strip() for row in feature_rows)
    assert all(row["Expected Behavior"] for row in feature_rows)


def test_qa_tracker_preserves_no_restart_and_fix_followup_state() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    overview = {str(row["Field"]): str(row["Value"]) for row in rows_by_header(wb["Overview"])}
    assert "Do not restart" in overview["No-restart constraint"]
    assert "Manual Retest Queue" in overview
    assert "Do not restart this active mycmux session" in overview["Next Safe Step"]

    errors = {str(row["Error ID"]): row for row in rows_by_header(wb["Error Log"])}
    assert errors["ERR-000"]["Status"] == "Open"
    assert "restart" in str(errors["ERR-000"]["Observed Behavior"]).lower()
    assert errors["ERR-003"]["Status"] == "Fixed in source; live retest deferred"
    assert errors["ERR-003"]["Retest Status"] == "Code retest pass; live runtime pending no-restart window"


def test_manual_retest_queue_covers_all_pending_or_deferred_stories() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    feature_rows = rows_by_header(wb["Feature Status"])
    matrix_rows = {str(row["Story ID"]): row for row in rows_by_header(wb["Test Matrix"])}
    queue_rows = rows_by_header(wb["Manual Retest Queue"])
    queue_ids = {str(row["Story ID"]) for row in queue_rows}

    expected_ids: set[str] = set()
    for story in feature_rows:
        story_id = str(story["Story ID"])
        matrix = matrix_rows[story_id]
        combined = "\n".join(
            str(value or "")
            for value in [story["Test Status"], story["Retest Status"], matrix["Status"], matrix["Current Result"]]
        ).lower()
        if "pending" in combined or "deferred" in combined or "backlog" in combined:
            expected_ids.add(story_id)

    assert expected_ids
    assert expected_ids == queue_ids
    assert {str(row["Safety Bucket"]) for row in queue_rows}.issuperset(REQUIRED_QUEUE_BUCKETS)
    assert all(str(row["Needs Safe Window"]) for row in queue_rows)
    assert all(str(row["Blocker / Gate"]) for row in queue_rows)


def test_manual_retest_queue_prioritizes_dangerous_live_checks() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    queue_rows = rows_by_header(wb["Manual Retest Queue"])
    by_story = {str(row["Story ID"]): row for row in queue_rows}

    assert by_story["MYC-050"]["Safety Bucket"] == "reload-required-live-fix-retest"
    assert by_story["MYC-001"]["Safety Bucket"] == "restart-or-duplicate-launch-window"
    assert by_story["MYC-002"]["Safety Bucket"] == "restart-or-duplicate-launch-window"
    assert by_story["MYC-003"]["Safety Bucket"] == "restart-or-duplicate-launch-window"
    assert by_story["MYC-044"]["Safety Bucket"] == "side-effect-gated"
    assert by_story["MYC-053"]["Safety Bucket"] == "backlog-confirmation"


def test_safe_window_playbook_matches_retest_queue() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    queue_rows = rows_by_header(wb["Manual Retest Queue"])
    playbook_rows = rows_by_header(wb["Safe Window Playbook"])

    queue_buckets = {str(row["Safety Bucket"]) for row in queue_rows}
    playbook_buckets = {str(row["Safety Bucket"]) for row in playbook_rows}
    assert playbook_buckets == queue_buckets

    queue_by_bucket: dict[str, set[str]] = {}
    for row in queue_rows:
        queue_by_bucket.setdefault(str(row["Safety Bucket"]), set()).add(str(row["Story ID"]))

    for row in playbook_rows:
        bucket = str(row["Safety Bucket"])
        story_ids = {story_id.strip() for story_id in str(row["Story IDs"]).split(",")}
        assert story_ids == queue_by_bucket[bucket]
        assert int(row["Story Count"]) == len(story_ids)
        assert "No restart" in str(row["No-Restart Guard"])
        assert "explicit approval" in str(row["No-Restart Guard"])
        assert str(row["Writeback Targets"]).startswith("Feature Status")

    by_bucket = {str(row["Safety Bucket"]): row for row in playbook_rows}
    assert by_bucket["reload-required-live-fix-retest"]["Phase"] == "Phase 1"
    assert "ERR-003" in str(by_bucket["reload-required-live-fix-retest"]["Actions"])
    assert "current no-restart constraint" in str(by_bucket["restart-or-duplicate-launch-window"]["Status"])
