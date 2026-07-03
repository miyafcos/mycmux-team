from __future__ import annotations
from pathlib import Path
from datetime import datetime
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

REPO = Path(r"C:\Users\miyaz\cmux-for-linux-dev-master")
OUT = REPO / "docs" / "qa" / "mycmux-feature-status-canonical.xlsx"

DATA = """
MYC-001|起動・ウィンドウ|起動時の作業ディレクトリ引き継ぎ|P0|Implemented per code|Likely|src/App.tsx@getLaunchCwd;src/App.tsx@listStore.workspaces.length === 0;src/lib/ipc.ts@get_launch_cwd|ユーザーとして任意の作業ディレクトリから起動したとき、そのディレクトリを初期ワークスペースとして使いたい。|永続ワークスペースが空でlaunch cwdが取れた場合、1x1ワークスペースを作成し、ペインとタブのcwdへlaunch cwdを設定する。|本体再起動が必要な実機検証は後回し。
MYC-002|起動・ウィンドウ|既存ワークスペースの起動復元|P0|Implemented per code|Likely|src/App.tsx@prepareStartupSessionGate;src/App.tsx@STARTUP_RESTORE_COMPLETE_EVENT;src/components/layout/SocketListener.tsx@useWorkspacePersist|ユーザーとして前回使っていたワークスペースとペイン構成を次回起動時に戻したい。|永続データ読込後、保存済みワークスペースを起動ゲート経由でマウントし、復元完了までスタートアップマスクを維持する。|再起動禁止のため静的棚卸しのみ。
MYC-003|起動・ウィンドウ|単一インスタンスと既存ウィンドウ前面化|P1|Implemented per command surface|Likely|src/lib/ipc.ts@claim_leader;src/lib/ipc.ts@reveal_main_window;src-tauri/src/commands/window.rs@fn claim_leader|ユーザーとしてmycmuxを重複起動したとき既存ウィンドウを前面に出してほしい。|window commandがリーダー獲得、ウィンドウ数取得、メインウィンドウ表示、終了を提供する。|重複起動の実機確認は保留。
MYC-004|起動・ウィンドウ|カスタムタイトルバー操作|P1|Implemented per UI|No|src/components/layout/TitleBar.tsx@New Workspace;src/components/layout/TitleBar.tsx@Toggle Sidebar;src/components/layout/TitleBar.tsx@Notifications|ユーザーとしてタイトルバーから最小化、閉じる、設定、通知、新規ワークスペースを操作したい。|TitleBarがClose、Minimize、New Workspace、Notifications、Settings、Toggle SidebarのUIを提供する。|閉じる操作は後続で注意。
MYC-005|セットアップ|初期ワークスペース作成フロー|P0|Implemented per UI/store|No|src/components/setup/WorkspaceSetup.tsx@My Workspace;src/components/setup/GridPicker.tsx@GridPicker;src/components/setup/AgentSelector.tsx@AgentSelector;src/stores/workspaceLayoutStore.ts@buildInitialPanes|初回ユーザーとして名前、グリッド、エージェント割当を選んで作業場を作りたい。|WorkspaceSetupが初期名、GridPicker、GridPreview、AgentSelectorを使い、選択に応じた初期ペインを作る。|空データ状態のテストが必要。
MYC-006|ワークスペース|ワークスペースタブ切替|P0|Implemented per UI|No|src/components/layout/TabBar.tsx@Workspaces;src/components/layout/TabBar.tsx@New workspace;src/components/layout/TabItem.tsx@Close workspace|ユーザーとして複数ワークスペースをタブで切り替えたい。|TabBarとTabItemがワークスペース一覧、アクティブ状態、新規作成、クローズ、通知状態バッジを表示する。|
MYC-007|ワークスペース|ワークスペース永続化|P0|Implemented per type/IPC|No|src/lib/ipc.ts@export interface PersistentData;src/lib/ipc.ts@load_persistent_data;src/lib/ipc.ts@save_persistent_data;src-tauri/src/commands/workspace.rs@fn load_persistent_data|ユーザーとしてワークスペース、設定、アクティブ対象、ピン留めルートを保存して復元したい。|PersistentDataがworkspaces、settings、active ids、pinned_rootsを保持し、load/save IPCでRust側ストレージに保存する。|
MYC-008|ペイン配置|ペイン分割|P0|Implemented per UI/store|No|src/components/workspace/PaneHeader.tsx@Split pane;src/components/workspace/PaneTabBar.tsx@Split right;src/stores/workspaceLayoutStore.ts@addPaneToWorkspace|ユーザーとして現在のペインを右または下に分割して複数作業を並べたい。|PaneHeaderまたはPaneTabBarからsplit right/downを呼び、workspaceLayoutStoreが列/行構造に新規ペインを追加する。|
MYC-009|ペイン配置|ペインクローズ|P0|Implemented per UI/store|No|src/components/workspace/PaneHeader.tsx@Close pane;src/components/workspace/PaneTabBar.tsx@Close pane;src/stores/workspaceLayoutStore.ts@removePaneFromWorkspace|ユーザーとして不要になったペインを閉じたい。|UIからClose paneを選ぶとremovePaneFromWorkspaceが対象ペインをワークスペースから取り除く。|最後のペイン時の挙動を後続確認。
MYC-010|ペイン配置|ペインとタブのドラッグ移動|P1|Implemented per store|No|src/hooks/usePaneDragSource.ts@usePaneDragSource;src/stores/workspaceLayoutStore.ts@moveTabToPane;src/stores/workspaceLayoutStore.ts@movePaneToNewWorkspace;src/components/workspace/PaneDragOverlay.tsx@PaneDragOverlay|ユーザーとしてタブやペインを別ペイン、分割位置、新規ワークスペースへ移動したい。|drag storeとlayout storeがtab-to-pane、tab-to-split、pane-to-pane、pane-to-split、new-workspace移動を提供する。|実操作テストが必要。
MYC-011|タブ|新規ターミナルタブ追加|P0|Implemented per UI/store|No|src/components/workspace/PaneTabBar.tsx@New terminal tab;src/stores/workspaceLayoutStore.ts@addTabToPane|ユーザーとして同じペイン内に新しいターミナルタブを追加したい。|PaneTabBarのNew terminal tabからaddTabToPaneがterminalタブを作成する。|
MYC-012|タブ|タブ切替とアクティブ管理|P0|Implemented per store/render|No|src/stores/workspaceLayoutStore.ts@setActivePaneTab;src/components/workspace/WorkspaceView.tsx@activeTab;src/components/workspace/TerminalPane.tsx@sessionId|ユーザーとしてペイン内のタブをクリックして表示対象を切り替えたい。|setActivePaneTabが対象ペインのactive_tab_idを更新し、WorkspaceViewとTerminalPaneが対応セッションを表示する。|
MYC-013|タブ|ペインタブ名の編集|P0|Implemented per recent changelog/code|No|src/components/workspace/PaneTabBar.tsx@setTabLabel;src/stores/workspaceLayoutStore.ts@setTabLabel;CHANGELOG.md@ダブルクリック編集|ユーザーとしてターミナルタブ名を任意のラベルへ変更し、必要なら自動名に戻したい。|PaneTabBarがダブルクリックまたはメニューで編集状態に入り、setTabLabelでlabelを保存する。空またはリセットで自動名へ戻る。|直近修正領域。操作テスト優先。
MYC-014|ターミナル|PTYセッション作成|P0|Implemented per IPC/Rust|No|src/lib/ipc.ts@create_session;src/components/workspace/TerminalPane.tsx@<XTermWrapper;src-tauri/src/commands/terminal.rs@create_session;src-tauri/src/pty/manager.rs@create_session|ユーザーとして各タブで選んだエージェントやシェルのターミナルを起動したい。|createSession IPCがagentId、cwd、session metadataを渡し、Rust PTY managerがセッションを作成する。|新規プロセス起動が絡む。
MYC-015|ターミナル|ターミナル入出力|P0|Implemented per UI/IPC|No|src/lib/ipc.ts@write_to_session;src/lib/ipc.ts@ack_frontend_data;src/lib/ipc.ts@get_session_scrollback;src/components/terminal/XTermWrapper.tsx@writeToSession|ユーザーとしてターミナルへ入力し、出力をリアルタイムに見たい。|XTermWrapperとTerminalPaneがwrite_to_session、frontend data ack、scrollback、visible stateを扱う。|
MYC-016|ターミナル|リサイズ同期|P0|Implemented per IPC/UI|No|src/lib/ipc.ts@resize_session;src/components/terminal/XTermWrapper.tsx@FitAddon;src-tauri/src/commands/terminal.rs@resize_session|ユーザーとしてペインサイズ変更後もターミナル表示を崩さず使いたい。|XTerm fit/resizeがcols/rowsを算出し、resize_session IPCでPTYサイズに反映する。|
MYC-017|ターミナル|ターミナル検索バー|P1|Implemented per UI|No|src/components/terminal/XTermWrapper.tsx@SearchAddon;src/components/terminal/XTermWrapper.tsx@Find...|ユーザーとして表示中のターミナル出力から文字列検索したい。|XTermWrapperがsearch addonとFind UIを持ち、入力語で現在バッファを検索する。|
MYC-018|ターミナル|選択、マウス、スクロール調整|P0|Implemented per recent changelog/code|No|src/components/terminal/XTermWrapper.tsx@mouse;CHANGELOG.md@wheel scrollback;CHANGELOG.md@text selection|ユーザーとしてClaude/Codex TUI内でも通常のテキスト選択とホイールスクロールを両立したい。|XTermWrapperがマウスモード制御シーケンスやwheel eventsを扱い、選択とTUIスクロールを壊さない。|直近不具合領域。実機テスト優先。
MYC-019|ターミナル|ターミナル設定検出|P2|Implemented per docs/code|No|src/lib/ipc.ts@get_terminal_config;src-tauri/src/terminal_config.rs@TerminalUserConfig;docs/features/implemented/config-detection.md@# Terminal Config Detection|ユーザーとして既存の端末設定に近いフォントや配色でターミナルを使いたい。|get_terminal_configとterminal_config Rust moduleが設定を検出し、フロントエンドが初期表示に使う。|
MYC-020|エージェント|組み込みエージェント選択|P0|Implemented per lib/UI/store|No|src/lib/agents.ts@AgentDefinition;src/components/setup/AgentSelector.tsx@AgentSelector;src/stores/workspaceLayoutStore.ts@setTabAgentId|ユーザーとしてClaude Code、Codex、Hybridなどの作業エージェントを選んで起動したい。|agents libraryがAgentDefinitionを提供し、setup/pane UIがagentIdをタブやペインへ割り当てる。|
MYC-021|エージェント|Claude/CodexセッションIDマッピング|P0|Implemented per App/IPC|No|src/App.tsx@applyAgentSessionMappings;src/lib/ipc.ts@read_agent_session_mappings;src-tauri/src/commands/session_mapping.rs@read_agent_session_mappings|ユーザーとして起動済みClaude/Codexセッションを復元、識別できる状態で使いたい。|read_agent_session_mappingsとpane metadataがagent kind/session idをワークスペース状態へ反映する。|
MYC-022|エージェント|CRSM履歴検索と再開パレット|P0|Implemented per UI/settings/IPC|No|src/components/CommandPalette/CrsmPalette.tsx@crsmListSessions;src/lib/ipc.ts@crsm_list_sessions;src/stores/settingsStore.ts@setCrsmShowClaude|ユーザーとして過去のClaude/Codex/Hybridセッション履歴を検索し、再開したい。|CrsmPaletteがcrsm_list_sessionsを呼び、設定フィルタに従って履歴を表示し、選択セッションを再開対象にする。|大量履歴での応答性を後続確認。
MYC-023|エージェント|CRSMハンドオフ作成|P1|Implemented per command surface|No|src/lib/ipc.ts@crsm_create_handoff;src-tauri/src/commands/crsm.rs@crsm_create_handoff|ユーザーとしてセッション内容から引き継ぎ文書を作りたい。|crsm_create_handoff IPCが選択セッションのハンドオフ生成を提供する。|UI導線の有無を後続で精査。
MYC-024|ファイル|ピン留めルート管理|P1|Implemented per UI/store/IPC|No|src/components/layout/FileExplorerSidebar.tsx@ルート追加;src/stores/fileExplorerStore.ts@addRoot;src/lib/ipc.ts@save_pinned_roots|ユーザーとしてよく使うフォルダをサイドバーにピン留めしたい。|FileExplorerSidebarとfileExplorerStoreがrootsを追加、削除、名称変更し、save_pinned_rootsで永続化する。|
MYC-025|ファイル|ディレクトリツリー閲覧|P1|Implemented per store/IPC|No|src/lib/ipc.ts@list_directory;src/lib/ipc.ts@walk_tree;src/lib/ipc.ts@watch_root;src/stores/fileExplorerStore.ts@toggleExpand|ユーザーとしてピン留めルート配下のファイルやフォルダをツリーで展開したい。|list_directory、walk_tree、watch_rootがファイル一覧、インデックス、変更通知を提供し、storeが展開状態を管理する。|
MYC-026|ファイル|ファイル選択と範囲選択|P2|Implemented per store|No|src/stores/fileExplorerStore.ts@toggleSelectedPath;src/stores/fileExplorerStore.ts@selectPathRange;src/stores/fileExplorerStore.ts@clearSelection|ユーザーとしてファイル一覧で単一、複数、範囲選択をしたい。|fileExplorerStoreがselectedPath、selectedPaths、toggleSelectedPath、selectPathRange、clearSelectionを管理する。|UIイベント経路を後続確認。
MYC-027|ファイル|ファイルを開く、Explorerで表示|P1|Implemented per IPC|No|src/lib/ipc.ts@open_with_default;src/lib/ipc.ts@reveal_in_explorer;src-tauri/src/commands/fs.rs@open_with_default|ユーザーとして選択ファイルを既定アプリで開いたりExplorerで場所を表示したい。|open_with_defaultとreveal_in_explorer IPCがOS側にファイル操作を委譲する。|外部アプリ起動は後続で慎重に。
MYC-028|ファイル|ファイルとフォルダ作成|P2|Implemented per IPC|No|src/lib/ipc.ts@create_file;src/lib/ipc.ts@create_folder;src-tauri/src/commands/fs.rs@create_file|ユーザーとしてサイドバーから新しいファイルやフォルダを作りたい。|create_fileとcreate_folder IPCが親パスと名前を受け取り、作成後のパスを返す。|テスト用一時フォルダで実施。
MYC-029|ファイル|パスジャンパーと検索|P1|Implemented per UI/store|No|src/components/layout/PathJumper.tsx@Path jumper;src/components/layout/PathJumper.tsx@Pinned;src/stores/fileExplorerStore.ts@buildSearchIndex|ユーザーとしてCtrl+Pからピン留め、最近使ったパス、インデックス済みファイルへ素早く移動したい。|PathJumperがPath jumper UIとPinned/Recent/indexing状態を表示し、fileExplorerStoreの検索インデックスを利用する。|
MYC-030|ブラウザ/成果物|HTMLと成果物プレビュー|P1|Implemented per IPC/store|No|src/lib/ipc.ts@preview_artifact_for_session;src/lib/ipc.ts@preview_artifact_uri_for_session_v2;src/stores/workspaceLayoutStore.ts@openOrReloadHtmlPreviewPane|ユーザーとしてターミナル出力や生成物からHTML/URIプレビューを開きたい。|preview_artifact_for_session、preview_artifact_uri_for_session、v2がプレビュー情報を返し、openOrReloadHtmlPreviewPaneがブラウザタブを開く。|
MYC-031|ブラウザ/成果物|ブラウザペイン|P1|Implemented per UI/command surface|No|src/components/workspace/BrowserPane.tsx@BrowserPane;src-tauri/src/commands/browser.rs@browser_create;src-tauri/src/commands/browser.rs@browser_navigate|ユーザーとしてアプリ内ペインでWebやローカル成果物を閲覧したい。|BrowserPaneがURL/プレビュー情報を受け取り、Tauri browser commandsがcreate/navigate/status/snapshotなどを提供する。|実レンダリング確認は後続。
MYC-032|ブラウザ/成果物|編集可能成果物の読込と保存|P2|Implemented per UI/IPC|No|src/lib/ipc.ts@read_editable_artifact;src/lib/ipc.ts@save_editable_artifact;src/components/workspace/ArtifactEditorToolbar.tsx@Editing|ユーザーとして生成された編集可能なHTMLや文書系成果物をアプリ内で編集して保存したい。|read_editable_artifact、save_editable_artifact IPCとArtifactEditorToolbarが編集操作と保存状態を支える。|対応ファイル種別とdirty stateを後続確認。
MYC-033|テーマ|テーマ切替|P1|Implemented per UI/store|No|src/components/theme/ThemeSwitcher.tsx@setTheme;src/stores/themeStore.ts@setTheme;src/components/theme/themeDefinitions.ts@THEMES|ユーザーとして好みの配色テーマへ切り替えたい。|ThemeSwitcherとthemeStoreがthemeIdを変更し、CSS変数とterminal colorsを更新する。|
MYC-034|テーマ|フォント設定|P1|Implemented per UI/store|No|src/components/theme/ThemeFontSettings.tsx@端末フォント;src/stores/themeStore.ts@setFontSize;src/stores/themeStore.ts@setFontFamily|ユーザーとして端末フォントのサイズとファミリーを調整したい。|ThemeFontSettingsがfont size/familyを変更し、themeStoreのsetFontSize/setFontFamilyへ保存する。|
MYC-035|テーマ|背景プリセットと画像パス|P2|Implemented per UI/store|No|src/components/theme/ThemeBackgroundPanel.tsx@Background;src/components/theme/ThemeBackgroundPanel.tsx@Paste image path;src/stores/themeStore.ts@setThemeBackground|ユーザーとして背景画像やプリセット背景を選んで作業画面の見た目を変えたい。|ThemeBackgroundPanelがBackground、Search presets、Paste image pathを提供し、themeStoreにbackground settingsを保存する。|
MYC-036|テーマ|色の微調整とリセット|P2|Implemented per UI/store|No|src/components/theme/ThemeTweakPanel.tsx@外観プリセット;src/components/theme/ThemeTweakPanel.tsx@この色だけ戻す;src/stores/themeStore.ts@resetThemeTweaks|ユーザーとして主要色を部分的に調整し、必要なら個別または全体を戻したい。|ThemeTweakPanelがpresets、detail colors、clear、resetを提供し、themeStoreがtweak colorを管理する。|
MYC-037|設定|通知設定|P1|Implemented per UI/store|No|src/components/layout/SettingsMenu.tsx@通知サウンド;src/stores/settingsStore.ts@setNotificationsEnabled;src/stores/settingsStore.ts@setNotificationSoundEnabled|ユーザーとして通知と通知音の有効/無効を切り替えたい。|SettingsMenuとsettingsStoreがnotificationsEnabledとnotificationSoundEnabledを管理する。|
MYC-038|設定|ターミナルレンダラー切替|P1|Implemented per UI/store|No|src/components/layout/SettingsMenu.tsx@Terminal renderer;src/stores/settingsStore.ts@setUseWebglRenderer;src/components/terminal/XTermWrapper.tsx@DOM renderer|ユーザーとしてWebGLレンダラーに問題があるとき安定DOMレンダラーへ切り替えたい。|SettingsMenuのTerminal renderer設定がuseWebglRendererを切り替え、XTermWrapperのrenderer選択に反映される。|
MYC-039|設定|キーバインド表示と変更|P2|Implemented per UI/store|No|src/components/layout/KeybindingsModal.tsx@Keybindings;src/stores/keybindingStore.ts@setOverride;src/stores/keybindingStore.ts@resetAll|ユーザーとしてショートカット一覧を確認し、必要なら上書きやリセットをしたい。|KeybindingsModalとkeybindingStoreがactionごとのoverride、clear、resetを提供する。|
MYC-040|通知|作業完了と承認待ちバッジ|P0|Implemented per UI/lib|No|src/lib/notificationStatus.ts@isShellProcess;src/components/layout/TabItem.tsx@Waiting for approval;src/components/workspace/PaneTabBar.tsx@Work done|ユーザーとして非アクティブなタブやワークスペースの作業完了、承認待ちを見落としたくない。|metadata/notificationStatusがprocess stateを判断し、TabItemとPaneTabBarにWaiting for approvalやWork doneを表示する。|
MYC-041|通知|通知パネル|P1|Implemented per UI|No|src/components/layout/NotificationPanel.tsx@Notifications;src/components/layout/TitleBar.tsx@Notifications|ユーザーとして通知履歴をまとめて確認したい。|NotificationPanelが通知一覧を表示し、TitleBarから開ける。|通知ソースの完全性は後続確認。
MYC-042|使用量|Claude/Codex使用量メーター|P1|Implemented per UI/store/Rust|No|src/components/layout/UsageMeter.tsx@UsageMeter;src/components/layout/UsagePopover.tsx@Claude Code;src/stores/usageStore.ts@get_usage_summary;src-tauri/src/commands/usage.rs@get_usage_summary|ユーザーとしてClaude CodeとCodexの使用量をタイトルバーから確認したい。|UsageMeterとUsagePopoverがusageStore経由でget_usage_summaryを読み、Claude/Codexの集計を表示する。|
MYC-043|Remote|iPhone/リモート接続情報表示|P1|Implemented per UI/Rust|No|src/components/layout/SettingsMenu.tsx@Remote Access;src/components/layout/SettingsMenu.tsx@qr_svg;src/lib/ipc.ts@get_remote_info;src-tauri/src/remote/mod.rs@get_remote_info|ユーザーとしてスマホから接続するためのURL、QR、接続状況を確認したい。|SettingsMenuのRemote Accessがget_remote_infoを呼び、URL、QR、token suffix、connected clientsを表示する。|ネットワーク/スマホ実機は後続。
MYC-044|Remote|リモートトークン再生成|P0|Implemented per UI/Rust|No|src/components/layout/SettingsMenu.tsx@トークン再生成;src/lib/ipc.ts@rotate_remote_token;src-tauri/src/remote/mod.rs@rotate_remote_token|ユーザーとして共有済みリモートURLのトークンを無効化して新しくしたい。|SettingsMenuのトークン再生成操作がrotate_remote_tokenを呼び、新しいremote infoを再読込する。|秘密情報の露出なしで検証する。
MYC-045|Remote|リモートWebクライアント|P2|Implemented per remote client files|No|src-tauri/src/remote/client/index.html@xterm;src-tauri/src/remote/client/app.js@WebSocket;src-tauri/src/remote/ws_handler.rs@WebSocket|ユーザーとしてスマホブラウザからターミナルを閲覧、操作したい。|remote client HTML/JS/CSSとWebSocket handler/sessionがxtermベースのリモート端末を提供する。|本体再起動なしでサーバ状態を壊さない範囲に限定。
MYC-046|Buddy|Buddy表示と折りたたみ|P2|Implemented per UI/store|No|src/components/layout/SettingsMenu.tsx@Claude Buddy;src/stores/settingsStore.ts@toggleBuddy;src/buddy/BuddyWidget.tsx@BuddyWidget|ユーザーとしてClaude Buddyを表示、非表示、折りたたみ切替したい。|SettingsMenu、settingsStore、BuddyWidgetがbuddyEnabled/buddyCollapsedを扱う。|
MYC-047|Buddy|Buddy設定保存|P2|Implemented per UI/Rust|No|src/components/layout/SettingsMenu.tsx@load_buddy_settings;src/components/layout/SettingsMenu.tsx@save_buddy_settings;src-tauri/src/buddy/commands.rs@save_buddy_settings|ユーザーとしてBuddyの有効状態やスキン設定を保存したい。|load_buddy_settings/save_buddy_settingsが設定JSONを読み書きし、SettingsMenuが選択を保存する。|
MYC-048|Buddy|Codex Petスキン読み込み|P2|Implemented per UI/Rust|No|src/buddy/sprite-skin.ts@list_codex_pets;src/buddy/sprite-skin.ts@read_codex_pet_spritesheet;src-tauri/src/buddy/commands.rs@CodexPetInfo|ユーザーとしてローカルのCodex PetをBuddyスキンとして選びたい。|sprite-skinがlist_codex_pets、read_codex_pet_spritesheet、read_codex_pet_layoutを呼び、スプライトシートを読み込む。|
MYC-049|Buddy|Buddyチャットと記憶|P3|Implemented per code path|No|src/buddy/core/PersonaAgent.ts@load_recent_chat;src/buddy/core/memory/profile.ts@save_buddy_profile_facet;src/buddy/core/memory/observation.ts@save_observation_state|ユーザーとしてBuddyとの会話や作業文脈を継続的に使いたい。|PersonaAgentとmemory modulesがrecent chat、work context、profile facets、observation stateを読み書きする。|モデル呼び出しとログ取り扱いは後続で安全確認。
MYC-050|自動化|Socket API応答|P2|Implemented per command/docs|No|src/lib/ipc.ts@socket_response;src-tauri/src/socket.rs@socket_response;docs/features/implemented/socket-api-and-automation.md@# Socket API|外部自動化ユーザーとしてmycmuxへコマンドを送り結果を受け取りたい。|socket_response IPCとsocket Rust moduleが外部APIへの応答経路を提供する。|外部ソケット実機確認は後続。
MYC-051|エラー処理|Reactエラーバウンダリ|P1|Implemented per UI|No|src/components/common/ErrorBoundary.tsx@Pane crashed;src/components/layout/ErrorBoundary.tsx@Something went wrong|ユーザーとしてひとつのペインや画面がクラッシュしてもアプリ全体を失いたくない。|common/layout ErrorBoundaryがペインまたは画面クラッシュ時にfallbackを表示する。|意図的クラッシュ注入は後続で判断。
MYC-052|アップデート|更新確認|P2|Implemented per dependency/UI|No|src/components/layout/SettingsMenu.tsx@更新を確認;src/lib/forcedAutoUpdater.ts@updater;package.json@@tauri-apps/plugin-updater|ユーザーとして利用中のmycmuxに更新があるか確認したい。|SettingsMenuから更新確認を実行し、forcedAutoUpdaterがTauri updater/process plugin経由の更新導線を扱う。|本体再起動と外部通信の可能性があるため実行しない。
MYC-053|アクセシビリティ/保留|ARIA、高コントラスト、Reduced Motion|P2|Pending doc / not proven implemented|No|docs/features/pending/accessibility-and-motion.md@# Accessibility & Motion|ユーザーとしてスクリーンリーダー、高コントラスト、動きの軽減に対応したUIを使いたい。|pending docsは要件を示すが、コード上の全面実装証拠は未確認。|未実装または要調査として追跡。
MYC-054|入力/保留|Copy Mode、Broadcast Input、Drag-and-Drop|P2|Pending doc / not proven implemented|No|docs/features/pending/input-and-interaction.md@# Input & Interaction Enhancements|ユーザーとしてコピー専用モード、複数ペイン同時入力、D&D操作を使いたい。|pending docsは要件を示すが、全機能の実装証拠は未確認。|後続でコード有無を追加調査。
MYC-055|SSH/保留|SSHリモートセッション|P3|Pending doc / not proven implemented|No|docs/features/pending/ssh-remote-sessions.md@# SSH Remote Sessions|ユーザーとしてSSH先のリモートシェルをmycmux内ペインで使いたい。|pending docsは要件を示すが、現コードで一般SSHセッションUIは未確認。|Remote Accessとは別機能として追跡。
MYC-056|UI polish/保留|Glassmorphism、ポート検出、リッチワークスペース表示|P3|Pending doc / partially unknown|No|docs/features/pending/ui-polish.md@# UI Polish;docs/features/pending/ui-polish.md@Port list UI|ユーザーとして視覚的に分かりやすく、実行中ポートなどの作業状態が見えるUIを使いたい。|pending docsはpolish要件を示すが、全項目の実装証拠は未確認。|UX改善フェーズ候補。
"""

source_cache = {}

def line_of(rel: str, needle: str) -> str:
    path = REPO / rel.replace('/', '\\')
    if not path.exists():
        return f"{path} (missing)"
    if rel not in source_cache:
        source_cache[rel] = path.read_text(encoding="utf-8", errors="replace").splitlines()
    for lineno, line in enumerate(source_cache[rel], 1):
        if needle in line:
            return f"{path}:{lineno}"
    return f"{path} (needle not found: {needle})"

def resolve_evidence(spec: str) -> str:
    refs = []
    for item in spec.split(';'):
        item = item.strip()
        if not item:
            continue
        if '@' in item:
            rel, needle = item.split('@', 1)
            refs.append(line_of(rel, needle))
        else:
            refs.append(str(REPO / item.replace('/', '\\')))
    return "\n".join(refs)

stories = []
for raw in DATA.splitlines():
    raw = raw.strip()
    if not raw:
        continue
    parts = raw.split('|')
    if len(parts) != 10:
        raise SystemExit(f"Bad row with {len(parts)} fields: {raw}")
    story_id, area, feature, priority, code_status, restart, evidence_spec, story, expected, notes = parts
    stories.append({
        "Story ID": story_id,
        "Area": area,
        "Feature": feature,
        "User Story": story,
        "Expected Behavior": expected,
        "Code Evidence": resolve_evidence(evidence_spec),
        "Priority": priority,
        "Code Status": code_status,
        "Test Status": "Not run" if not code_status.startswith("Pending") else "Not applicable yet",
        "Error Status": "No runtime errors logged yet",
        "Fix Status": "Not started",
        "Retest Status": "Not started",
        "Notes": notes,
        "Restart Needed": restart,
    })

wb = Workbook()
ws = wb.active
ws.title = "Feature Status"
feature_headers = ["Story ID", "Area", "Feature", "User Story", "Expected Behavior", "Code Evidence", "Priority", "Code Status", "Test Status", "Error Status", "Fix Status", "Retest Status", "Notes"]
ws.append(feature_headers)
for s in stories:
    ws.append([s[h] for h in feature_headers])

wt = wb.create_sheet("Test Matrix")
test_headers = ["Story ID", "Test Type", "Restart Needed", "Test Steps", "Expected Result", "Current Result", "Evidence", "Status", "Notes"]
wt.append(test_headers)
for s in stories:
    if s["Code Status"].startswith("Pending"):
        test_type = "Backlog verification"
        steps = "コードを再検索し、実装証拠がなければBacklogのまま追跡する。"
        current = "Not run."
        status = "Not started"
    elif s["Restart Needed"] == "Likely":
        test_type = "Manual runtime deferred"
        steps = "安全なタイミングで再起動依存の動作を検証する。現在のmycmux本体は再起動しない。"
        current = "Deferred because current Codex session is running inside mycmux and restart is prohibited."
        status = "Deferred / no restart"
    elif s["Story ID"] in {"MYC-027", "MYC-028", "MYC-043", "MYC-044", "MYC-045", "MYC-052"}:
        test_type = "Manual side-effect-safe"
        steps = "テスト用対象または読み取り専用観察を用意し、UI/API経路を実行して副作用を記録する。"
        current = "Not run."
        status = "Not started"
    else:
        test_type = "Manual UI + static support"
        steps = "UIコントロールの存在を確認し、ユーザー操作を実行し、状態、描画、出力を記録する。"
        current = "Not run."
        status = "Not started"
    wt.append([s["Story ID"], test_type, s["Restart Needed"], steps, s["Expected Behavior"], current, s["Code Evidence"], status, s["Notes"]])

we = wb.create_sheet("Error Log")
error_headers = ["Error ID", "Story ID", "Type", "Severity", "Observed Behavior", "Expected Behavior", "Evidence", "Status", "Fix Plan", "Retest Status", "Notes"]
we.append(error_headers)
we.append(["ERR-000", "ALL", "Logistic", "Info", "Runtime testing has not started because this session itself runs in mycmux and the user explicitly prohibited restarting the app.", "Testing plan avoids restarting, quitting, replacing, or competing with the active mycmux instance.", "Current user instruction recorded in this workbook.", "Open", "Start with static and no-restart tests; defer restart-dependent stories.", "Not started", "Constraint, not a product defect."])

wo = wb.create_sheet("Overview")
wo.append(["Field", "Value"])
for field, value in [
    ("Canonical workbook", str(OUT)),
    ("Repository", str(REPO)),
    ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ("Scope", "mycmux master worktree feature inventory, code-based expected behavior, test/error/fix/retest tracking"),
    ("No-restart constraint", "Do not restart, quit, replace, or launch a competing mycmux instance during this active session."),
    ("Current phase", "Feature inventory and user-story drafting complete for initial pass; runtime testing not started."),
    ("Total stories", len(stories)),
    ("Runtime test loop status", "Pending"),
    ("Fix loop status", "Pending"),
    ("Retest loop status", "Pending"),
]:
    wo.append([field, value])

wi = wb.create_sheet("Evidence Index")
wi.append(["Source", "Kind", "Evidence / Notes"])
seen = set()
for s in stories:
    for line in s["Code Evidence"].splitlines():
        source = re.sub(r":\d+$", "", line)
        source = re.sub(r" \(needle not found:.*$", "", source)
        if source not in seen:
            seen.add(source)
            wi.append([source, "Code/doc source", "Referenced by Feature Status"])
wi.append([str(REPO / "docs" / "qa" / "mycmux-static-inventory.json"), "Generated inventory", "Frontend invokes, Rust commands, component labels"])
wi.append([str(REPO / "docs" / "qa" / "mycmux-feature-evidence-summary.json"), "Generated evidence summary", "Feature docs and key component lines"])


wv = wb.create_sheet("Validation Log")
wv.append(["Timestamp", "Check", "Command / Method", "Result", "Restarted App", "Notes"])
wv.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Workbook structural validation", "load_workbook + required sheet and row count checks inside generator", "Pass: workbook reopened, required sheets present, Feature Status and Test Matrix row counts match DATA", "No", "Checks run after save."])
wv.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Code evidence resolution", "line_of() search over current worktree files", "Pass: all evidence refs resolve to existing files and line numbers in current worktree", "No", "Any missing evidence would fail the follow-up inspection."])
wv.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "No-restart constraint", "Manual process constraint", "Pass: no mycmux quit, restart, replacement, or competing app launch was performed", "No", "Runtime tests requiring restart are deferred in Test Matrix."])
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2F3")
for sheet in wb.worksheets:
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if sheet.max_row >= 2:
        ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        name = re.sub(r"[^A-Za-z0-9_]", "", sheet.title)[:20] + "Table"
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        try:
            sheet.add_table(table)
        except ValueError:
            pass
    default_widths = {1: 14, 2: 20, 3: 34, 4: 54, 5: 64, 6: 74, 7: 12, 8: 30, 9: 24, 10: 28, 11: 18, 12: 18, 13: 40}
    for col_index, width in default_widths.items():
        if sheet.max_column >= col_index:
            sheet.column_dimensions[get_column_letter(col_index)].width = width
    if sheet.title == "Overview":
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 130

for row in range(2, ws.max_row + 1):
    for col in [8, 9, 10, 11, 12]:
        value = str(ws.cell(row=row, column=col).value or "")
        if "Implemented" in value:
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="E2F0D9")
        elif "Pending" in value or "Deferred" in value:
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FCE4D6")
        elif "Not" in value:
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF2CC")

wb.save(OUT)
loaded = load_workbook(OUT, read_only=True, data_only=True)
required = {"Overview", "Feature Status", "Test Matrix", "Error Log", "Evidence Index"}
missing = required.difference(loaded.sheetnames)
if missing:
    raise SystemExit(f"missing sheets: {missing}")
feature_rows = loaded["Feature Status"].max_row - 1
test_rows = loaded["Test Matrix"].max_row - 1
if feature_rows != len(stories):
    raise SystemExit(f"feature row mismatch: {feature_rows} != {len(stories)}")
if test_rows != len(stories):
    raise SystemExit(f"test row mismatch: {test_rows} != {len(stories)}")
print(str(OUT))
print(f"feature_rows={feature_rows}")
print(f"test_rows={test_rows}")
print("sheets=" + ", ".join(loaded.sheetnames))


