from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


REPO = Path(r"C:\Users\miyaz\cmux-for-linux-dev-master")
WORKBOOK = REPO / "docs" / "qa" / "mycmux-feature-status-canonical.xlsx"
LOG_DIR = REPO / "docs" / "qa" / "logs"
TIMESTAMP = "2026-06-25 02:10:00"

FINAL_LOGS = {
    "pytest": LOG_DIR / "validation-20260625-021000-pytest-final-followup.log",
    "tsc": LOG_DIR / "validation-20260625-021000-tsc-final-followup.log",
    "npm_build": LOG_DIR / "validation-20260625-021000-npm-build-final-followup.log",
    "cargo": LOG_DIR / "validation-20260625-021000-cargo-final-followup.log",
}

RUN = {
    "Run ID": "RUN-20260625-021000-FINAL-FOLLOWUP-VALIDATION",
    "Timestamp": TIMESTAMP,
    "Command": "python -m pytest tests -q; cmd /c npx tsc --noEmit; cmd /c npm run build; cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux",
    "Exit Code": 0,
    "Result Summary": "Pass: 41 Python tests, TypeScript noEmit, npm build, and 75 Rust tests passed; npm build kept existing chunk-size warning",
    "Restarted App": "No",
    "Log Path": "\n".join(str(path) for path in FINAL_LOGS.values()),
    "Covered Stories": "ALL source-contract and build-validation scope",
    "Notes": "Build/test validation only. No mycmux restart, quit, install, replacement, competing launch, token rotation, or live UI side effect was performed.",
}

VALIDATION_ROWS = [
    [TIMESTAMP, "Python full contract suite", "python -m pytest tests -q", "Pass: 41 passed in 0.08s", "No", str(FINAL_LOGS["pytest"])],
    [TIMESTAMP, "TypeScript noEmit final", "cmd /c npx tsc --noEmit", "Pass: exit code 0", "No", str(FINAL_LOGS["tsc"])],
    [TIMESTAMP, "Frontend production build final", "cmd /c npm run build", "Pass: build completed; existing chunk-size warning only", "No", str(FINAL_LOGS["npm_build"])],
    [TIMESTAMP, "Rust unit tests final", "cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux", "Pass: 75 passed; 0 failed", "No", str(FINAL_LOGS["cargo"])],
]


def headers(sheet) -> list[str]:
    return [str(cell.value) for cell in sheet[1]]


def row_by_key(sheet, key: str) -> dict[str, int]:
    hs = headers(sheet)
    col = hs.index(key) + 1
    return {str(sheet.cell(row=row, column=col).value): row for row in range(2, sheet.max_row + 1)}


def append_unique_row(sheet, key_col: str, values_by_header: dict[str, object]) -> None:
    hs = headers(sheet)
    rows = row_by_key(sheet, key_col)
    key = str(values_by_header[key_col])
    row = rows.get(key, sheet.max_row + 1)
    for col, header in enumerate(hs, start=1):
        if header in values_by_header:
            sheet.cell(row=row, column=col).value = values_by_header[header]


def append_validation(sheet, values: list[object]) -> None:
    existing = {(sheet.cell(row=row, column=2).value, sheet.cell(row=row, column=3).value) for row in range(2, sheet.max_row + 1)}
    key = (values[1], values[2])
    if key not in existing:
        sheet.append(values)


def append_evidence(sheet, source: str, kind: str, notes: str) -> None:
    existing = {str(sheet.cell(row=row, column=1).value) for row in range(2, sheet.max_row + 1)}
    if source not in existing:
        sheet.append([source, kind, notes])


def restyle_and_resize_tables(wb) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        if sheet.tables:
            for table in sheet.tables.values():
                table.ref = ref
        elif sheet.max_row >= 2:
            table_name = "".join(ch for ch in sheet.title if ch.isalnum())[:20] + "Table"
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            sheet.add_table(table)
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            if col == 1:
                sheet.column_dimensions[letter].width = 16
            elif col in (2, 7, 8, 9, 10, 11, 12):
                sheet.column_dimensions[letter].width = 24
            else:
                sheet.column_dimensions[letter].width = 42
        if sheet.title == "Overview":
            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 130


def main() -> None:
    missing_logs = [str(path) for path in FINAL_LOGS.values() if not path.exists()]
    if missing_logs:
        raise SystemExit(f"missing validation logs: {missing_logs}")

    wb = load_workbook(WORKBOOK)
    append_unique_row(wb["Test Run Log"], "Run ID", RUN)
    for row in VALIDATION_ROWS:
        append_validation(wb["Validation Log"], row)
    for key, path in FINAL_LOGS.items():
        append_evidence(wb["Evidence Index"], str(path), "Final validation log", f"Referenced by RUN-20260625-021000-FINAL-FOLLOWUP-VALIDATION ({key})")

    overview = wb["Overview"]
    overview_rows = row_by_key(overview, "Field")
    updates = {
        "Current phase": "No-restart source/runtime QA continued and final validation passed; live UI/restart-dependent retest remains open.",
        "Runtime test loop status": "In progress: source contracts, static checks, no-restart runtime checks, and full build/test validation pass. Manual UI, OS side effects, token rotation, duplicate launch, and restart restore checks remain deferred.",
        "Fix loop status": "ERR-003 source fix remains validated by tests/build; ERR-004 tracker cleanup fixed; no new product defect found in final validation.",
        "Retest loop status": "Passed: 41 Python tests, TypeScript noEmit, npm build, 75 Rust tests. Pending: live app reload/restart window for Socket API runtime retest and full user-behavior retest.",
    }
    for key, value in updates.items():
        row = overview_rows.get(key)
        if row:
            overview.cell(row=row, column=2).value = value

    restyle_and_resize_tables(wb)
    wb.save(WORKBOOK)
    loaded = load_workbook(WORKBOOK, read_only=True, data_only=False)
    for sheet_name in ["Feature Status", "Test Matrix", "Error Log", "Overview", "Evidence Index", "Validation Log", "Test Run Log"]:
        print(f"{sheet_name}: rows={loaded[sheet_name].max_row - 1}")


if __name__ == "__main__":
    main()

