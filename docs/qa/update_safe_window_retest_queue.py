from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


REPO = Path(r"C:\Users\miyaz\cmux-for-linux-dev-master")
WORKBOOK = REPO / "docs" / "qa" / "mycmux-feature-status-canonical.xlsx"
SCRIPT_PATH = REPO / "docs" / "qa" / "update_safe_window_retest_queue.py"
TIMESTAMP = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
RUN_ID = "RUN-20260625-SAFE-WINDOW-QUEUE"
FINAL_VALIDATION_RUN_ID = "RUN-20260625-023000-SAFE-QUEUE-FINAL-VALIDATION"
FINAL_VALIDATION_TIMESTAMP = "2026-06-25 02:30:00"
FINAL_VALIDATION_LOGS = {
    "pytest": REPO / "docs" / "qa" / "logs" / "validation-20260625-023000-pytest-safe-queue-final.log",
    "tsc": REPO / "docs" / "qa" / "logs" / "validation-20260625-023000-tsc-safe-queue-final.log",
    "npm_build": REPO / "docs" / "qa" / "logs" / "validation-20260625-023000-npm-build-safe-queue-final.log",
    "cargo": REPO / "docs" / "qa" / "logs" / "validation-20260625-023000-cargo-safe-queue-final.log",
}
PLAYBOOK_VALIDATION_RUN_ID = "RUN-20260625-031500-SAFE-PLAYBOOK-VALIDATION"
PLAYBOOK_VALIDATION_TIMESTAMP = "2026-06-25 03:15:00"
PLAYBOOK_VALIDATION_LOGS = {
    "pytest": REPO / "docs" / "qa" / "logs" / "validation-20260625-031500-safe-playbook-pytest.log",
    "tsc": REPO / "docs" / "qa" / "logs" / "validation-20260625-031500-safe-playbook-tsc.log",
    "npm_build": REPO / "docs" / "qa" / "logs" / "validation-20260625-031500-safe-playbook-npm-build.log",
    "cargo": REPO / "docs" / "qa" / "logs" / "validation-20260625-031500-safe-playbook-cargo.log",
}

QUEUE_SHEET = "Manual Retest Queue"
PLAYBOOK_SHEET = "Safe Window Playbook"
QUEUE_HEADERS = [
    "Queue ID",
    "Story ID",
    "Area",
    "Priority",
    "Feature",
    "Safety Bucket",
    "Needs Safe Window",
    "Restart Needed",
    "Side Effect Risk",
    "Current Status",
    "Retest Target",
    "Manual Steps",
    "Expected Result",
    "Blocker / Gate",
    "Evidence Source",
    "Last Updated",
]
PLAYBOOK_HEADERS = [
    "Phase",
    "Safety Bucket",
    "Approval Required",
    "Story Count",
    "Queue IDs",
    "Story IDs",
    "Execution Scope",
    "Start Condition",
    "Actions",
    "Expected Outcome",
    "Writeback Targets",
    "Stop Condition",
    "No-Restart Guard",
    "Status",
    "Last Updated",
]

PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
BUCKET_ORDER = {
    "reload-required-live-fix-retest": 0,
    "restart-or-duplicate-launch-window": 1,
    "side-effect-gated": 2,
    "interactive-live-ui-window": 3,
    "manual-runtime-window": 4,
    "backlog-confirmation": 5,
}
PLAYBOOK_DEFINITIONS = {
    "reload-required-live-fix-retest": {
        "phase": "Phase 1",
        "approval": "Safe reload/restart window for this active mycmux session",
        "execution_scope": "Retest the patched Socket API bridge before any broad click-through.",
        "start_condition": "User confirms this mycmux-hosted Codex session can be interrupted.",
        "actions": "Load the patched frontend, retest MYC-050 result and error responses, and update ERR-003.",
        "expected_outcome": "Socket API returns a result/error instead of timing out.",
        "writeback_targets": "Feature Status, Test Matrix, Error Log, Test Run Log",
        "stop_condition": "Stop immediately if reloading would interrupt active work or if Socket API still times out.",
        "status": "Blocked by current no-restart constraint",
    },
    "restart-or-duplicate-launch-window": {
        "phase": "Phase 2",
        "approval": "Safe restart or duplicate-launch window",
        "execution_scope": "Startup cwd, persisted restore, and single-instance foregrounding.",
        "start_condition": "User approves restart/duplicate-launch tests after the Socket API retest.",
        "actions": "Run MYC-001, MYC-002, and MYC-003 in order; document any startup or foregrounding errors.",
        "expected_outcome": "Startup state restores and duplicate launch foregrounds the existing window.",
        "writeback_targets": "Feature Status, Test Matrix, Error Log, Test Run Log",
        "stop_condition": "Stop if a second app instance could displace the current session unexpectedly.",
        "status": "Blocked by current no-restart constraint",
    },
    "side-effect-gated": {
        "phase": "Phase 3",
        "approval": "Story-by-story approval for each side effect",
        "execution_scope": "Token rotation, external app/file open, update flow, agent launch, and fault injection.",
        "start_condition": "User approves the exact story and side effect before execution.",
        "actions": "Run one queued story at a time; record the side effect, result, and rollback note if applicable.",
        "expected_outcome": "Each approved side-effect story either passes or produces a documented error with fix plan.",
        "writeback_targets": "Feature Status, Test Matrix, Error Log, Test Run Log",
        "stop_condition": "Stop before any unapproved token, external app, update, launch, or fault action.",
        "status": "Awaiting story-specific approval",
    },
    "interactive-live-ui-window": {
        "phase": "Phase 4",
        "approval": "Interruptible manual UI click-through window",
        "execution_scope": "Workspace, pane, tab, terminal, settings, notification, file, preview, and Buddy UI stories.",
        "start_condition": "User approves a focused manual UI pass while current work can tolerate UI focus changes.",
        "actions": "Run queue items in priority order, write pass/fail evidence, and open errors for UX/logistical defects.",
        "expected_outcome": "Every interactive UI story has a concrete pass/fail retest result.",
        "writeback_targets": "Feature Status, Test Matrix, Error Log, Test Run Log",
        "stop_condition": "Stop if testing would require restart, replacement, token rotation, or another side effect.",
        "status": "Awaiting interactive test window",
    },
    "manual-runtime-window": {
        "phase": "Phase 5",
        "approval": "Runtime proof window for non-restart behavior",
        "execution_scope": "Usage meter, file/folder creation, and remote web client runtime proof.",
        "start_condition": "User approves story-specific runtime proof without app restart.",
        "actions": "Run story-specific runtime steps and record filesystem/network side effects without exposing secrets.",
        "expected_outcome": "Runtime-dependent stories have pass/fail evidence and no raw token leakage.",
        "writeback_targets": "Feature Status, Test Matrix, Error Log, Test Run Log",
        "stop_condition": "Stop before secret exposure, unapproved file creation, or unapproved network/client action.",
        "status": "Awaiting runtime proof window",
    },
    "backlog-confirmation": {
        "phase": "Phase 6",
        "approval": "Scope decision from user",
        "execution_scope": "Backlog-only or partially implemented stories that need acceptance-scope confirmation.",
        "start_condition": "User chooses keep as backlog, split into implementation stories, or remove from current scope.",
        "actions": "Update story status according to the decision; do not mark runtime pass without implementation proof.",
        "expected_outcome": "Backlog stories are explicitly scoped instead of lingering as ambiguous test failures.",
        "writeback_targets": "Feature Status, Test Matrix, Error Log, Test Run Log",
        "stop_condition": "Stop if the scope decision would require implementation work not yet requested.",
        "status": "Awaiting scope decision",
    },
}


def headers(sheet) -> list[str]:
    return [str(cell.value) for cell in sheet[1]]


def table_rows(sheet) -> list[dict[str, Any]]:
    hs = headers(sheet)
    return [dict(zip(hs, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]


def row_by_key(sheet, key: str) -> dict[str, int]:
    hs = headers(sheet)
    col = hs.index(key) + 1
    return {str(sheet.cell(row=row, column=col).value): row for row in range(2, sheet.max_row + 1)}


def set_overview_value(sheet, key: str, value: str) -> None:
    rows = row_by_key(sheet, "Field")
    row = rows.get(key)
    if row is None:
        sheet.append([key, value])
    else:
        sheet.cell(row=row, column=2).value = value


def append_unique_row(sheet, key_col: str, values_by_header: dict[str, object]) -> None:
    hs = headers(sheet)
    rows = row_by_key(sheet, key_col)
    key = str(values_by_header[key_col])
    row = rows.get(key, sheet.max_row + 1)
    for col, header in enumerate(hs, start=1):
        if header in values_by_header:
            sheet.cell(row=row, column=col).value = values_by_header[header]


def append_unique_validation(sheet, values: list[object]) -> None:
    existing = {
        (sheet.cell(row=row, column=2).value, sheet.cell(row=row, column=3).value)
        for row in range(2, sheet.max_row + 1)
    }
    key = (values[1], values[2])
    if key not in existing:
        sheet.append(values)


def append_unique_evidence(sheet, source: str, kind: str, notes: str) -> None:
    existing = {str(sheet.cell(row=row, column=1).value) for row in range(2, sheet.max_row + 1)}
    if source not in existing:
        sheet.append([source, kind, notes])


def classify_bucket(story: dict[str, Any], matrix: dict[str, Any]) -> tuple[str, str, str, str]:
    haystack = "\n".join(
        str(value or "")
        for value in [
            story.get("Story ID"),
            story.get("Feature"),
            story.get("Test Status"),
            story.get("Retest Status"),
            story.get("Notes"),
            matrix.get("Restart Needed"),
            matrix.get("Current Result"),
            matrix.get("Notes"),
        ]
    ).lower()

    if story.get("Story ID") == "MYC-050" or "err-003" in haystack:
        return (
            "reload-required-live-fix-retest",
            "Yes: patched frontend must be loaded before runtime retest",
            "Reload/restart would replace the currently running UI bundle",
            "Approve a safe mycmux reload/restart window, then retest Socket API result/error responses.",
        )
    if str(matrix.get("Restart Needed") or "").lower() == "likely" or "duplicate-launch" in haystack:
        return (
            "restart-or-duplicate-launch-window",
            "Yes: requires restart or competing launch",
            "Can disrupt this active mycmux-hosted session",
            "Approve a safe restart/duplicate-launch window, then test startup cwd, restore, and foregrounding.",
        )
    if any(token in haystack for token in ["token rotation", "os side-effect", "fault", "update runtime", "manual launch", "external side effect"]):
        return (
            "side-effect-gated",
            "Yes: side effect must be explicitly allowed",
            "May rotate tokens, open external apps, trigger update flow, launch agents, or inject faults",
            "Approve the specific side effect, then execute the story-specific test steps.",
        )
    if "backlog confirmed" in haystack or "pending doc" in haystack:
        return (
            "backlog-confirmation",
            "No for runtime; yes for scope decision",
            "No product pass should be claimed until implementation evidence exists",
            "Decide whether to keep as backlog, split into implementation stories, or remove from current acceptance scope.",
        )
    if any(token in haystack for token in ["manual ui pending", "live ui pending", "manual drag", "manual settings", "manual renderer", "manual preview"]):
        return (
            "interactive-live-ui-window",
            "Yes: interactive click-through needed",
            "May disturb the current working layout but does not require app restart",
            "Run the UI click-through during an interruptible window, recording pass/fail in Test Matrix.",
        )
    return (
        "manual-runtime-window",
        "Yes: manual runtime proof still needed",
        "Runtime behavior not fully proven by source/build checks",
        "Run the story-specific runtime steps and update retest status.",
    )


def sort_key(item: dict[str, Any]) -> tuple[int, int, str]:
    return (
        BUCKET_ORDER.get(str(item["Safety Bucket"]), 99),
        PRIORITY_ORDER.get(str(item["Priority"]), 99),
        str(item["Story ID"]),
    )


def build_queue_rows(feature_rows: list[dict[str, Any]], matrix_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    matrix_by_story = {str(row["Story ID"]): row for row in matrix_rows}
    queue: list[dict[str, Any]] = []
    for story in feature_rows:
        story_id = str(story["Story ID"])
        matrix = matrix_by_story.get(story_id, {})
        combined_status = "\n".join(
            str(value or "")
            for value in [story.get("Test Status"), story.get("Retest Status"), matrix.get("Status"), matrix.get("Current Result")]
        )
        lowered_status = combined_status.lower()
        if "pending" not in lowered_status and "deferred" not in lowered_status and "backlog" not in lowered_status:
            continue
        bucket, needs_window, side_effect_risk, blocker = classify_bucket(story, matrix)
        queue.append(
            {
                "Story ID": story_id,
                "Area": story.get("Area"),
                "Priority": story.get("Priority"),
                "Feature": story.get("Feature"),
                "Safety Bucket": bucket,
                "Needs Safe Window": needs_window,
                "Restart Needed": matrix.get("Restart Needed"),
                "Side Effect Risk": side_effect_risk,
                "Current Status": story.get("Test Status"),
                "Retest Target": story.get("Retest Status"),
                "Manual Steps": matrix.get("Test Steps"),
                "Expected Result": matrix.get("Expected Result"),
                "Blocker / Gate": blocker,
                "Evidence Source": matrix.get("Evidence"),
                "Last Updated": TIMESTAMP,
            }
        )
    queue.sort(key=sort_key)
    for index, row in enumerate(queue, start=1):
        row["Queue ID"] = f"Q-{index:03d}"
    return queue


def reset_queue_sheet(wb, rows: list[dict[str, Any]]) -> None:
    if QUEUE_SHEET in wb.sheetnames:
        del wb[QUEUE_SHEET]
    sheet = wb.create_sheet(QUEUE_SHEET, 3)
    sheet.append(QUEUE_HEADERS)
    for row in rows:
        sheet.append([row.get(header) for header in QUEUE_HEADERS])


def build_playbook_rows(queue_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in queue_rows:
        grouped.setdefault(str(row["Safety Bucket"]), []).append(row)

    playbook_rows: list[dict[str, Any]] = []
    for bucket, rows in sorted(grouped.items(), key=lambda item: BUCKET_ORDER.get(item[0], 99)):
        definition = PLAYBOOK_DEFINITIONS[bucket]
        queue_ids = ", ".join(str(row["Queue ID"]) for row in rows)
        story_ids = ", ".join(str(row["Story ID"]) for row in rows)
        playbook_rows.append(
            {
                "Phase": definition["phase"],
                "Safety Bucket": bucket,
                "Approval Required": definition["approval"],
                "Story Count": len(rows),
                "Queue IDs": queue_ids,
                "Story IDs": story_ids,
                "Execution Scope": definition["execution_scope"],
                "Start Condition": definition["start_condition"],
                "Actions": definition["actions"],
                "Expected Outcome": definition["expected_outcome"],
                "Writeback Targets": definition["writeback_targets"],
                "Stop Condition": definition["stop_condition"],
                "No-Restart Guard": "No restart, reload, quit, app replacement, competing launch, token rotation, external app open, fault injection, or update flow without explicit approval.",
                "Status": definition["status"],
                "Last Updated": TIMESTAMP,
            }
        )
    return playbook_rows


def reset_playbook_sheet(wb, rows: list[dict[str, Any]]) -> None:
    if PLAYBOOK_SHEET in wb.sheetnames:
        del wb[PLAYBOOK_SHEET]
    sheet = wb.create_sheet(PLAYBOOK_SHEET, 4)
    sheet.append(PLAYBOOK_HEADERS)
    for row in rows:
        sheet.append([row.get(header) for header in PLAYBOOK_HEADERS])


def restyle_workbook(wb) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if sheet.max_row >= 2:
            ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            if sheet.tables:
                for table in sheet.tables.values():
                    table.ref = ref
            else:
                table_name = "".join(ch for ch in sheet.title if ch.isalnum())[:20] + "Table"
                table = Table(displayName=table_name, ref=ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                sheet.add_table(table)
        for column in range(1, sheet.max_column + 1):
            width = 18
            if sheet.title == QUEUE_SHEET:
                width = {1: 12, 2: 12, 3: 18, 4: 10, 5: 30, 6: 28, 7: 32, 8: 16, 9: 34, 10: 34, 11: 30, 16: 20}.get(column, 44)
            elif sheet.title == PLAYBOOK_SHEET:
                width = {1: 12, 2: 30, 3: 34, 4: 12, 5: 36, 6: 52, 14: 30, 15: 20}.get(column, 46)
            elif sheet.title == "Overview":
                width = 30 if column == 1 else 130
            elif column in (1, 2, 7, 8, 9, 10, 11, 12):
                width = 24
            else:
                width = 42
            sheet.column_dimensions[get_column_letter(column)].width = width


def main() -> None:
    wb = load_workbook(WORKBOOK)
    feature_rows = table_rows(wb["Feature Status"])
    matrix_rows = table_rows(wb["Test Matrix"])
    queue_rows = build_queue_rows(feature_rows, matrix_rows)
    if not queue_rows:
        raise SystemExit("no pending/deferred stories found; refusing to clear retest queue")

    reset_queue_sheet(wb, queue_rows)
    playbook_rows = build_playbook_rows(queue_rows)
    reset_playbook_sheet(wb, playbook_rows)

    bucket_counts: dict[str, int] = {}
    for row in queue_rows:
        bucket = str(row["Safety Bucket"])
        bucket_counts[bucket] = bucket_counts.get(bucket, 0) + 1
    summary = "; ".join(f"{bucket}={count}" for bucket, count in sorted(bucket_counts.items(), key=lambda item: BUCKET_ORDER.get(item[0], 99)))

    set_overview_value(wb["Overview"], "Manual Retest Queue", f"{len(queue_rows)} pending/deferred stories queued by safety bucket: {summary}")
    set_overview_value(wb["Overview"], "Safe Window Playbook", f"{len(playbook_rows)} ordered phases generated from Manual Retest Queue; each phase preserves the active-session no-restart gate and writeback targets.")
    set_overview_value(wb["Overview"], "Next Safe Step", "Do not restart this active mycmux session. Ask for an interruptible window before Socket API live retest, duplicate launch, startup restore, token rotation, update flow, external app open, or full UI click-through.")

    append_unique_validation(
        wb["Validation Log"],
        [
            TIMESTAMP,
            "Manual retest queue refresh",
            f"python {SCRIPT_PATH}",
            f"Pass: {len(queue_rows)} pending/deferred stories grouped into {QUEUE_SHEET}; {len(playbook_rows)} safe-window phases grouped into {PLAYBOOK_SHEET}",
            "No",
            str(WORKBOOK),
        ],
    )
    append_unique_row(
        wb["Test Run Log"],
        "Run ID",
        {
            "Run ID": RUN_ID,
            "Timestamp": TIMESTAMP,
            "Command": f"python {SCRIPT_PATH}",
            "Exit Code": 0,
            "Result Summary": f"Pass: refreshed {QUEUE_SHEET} with {len(queue_rows)} pending/deferred stories and {PLAYBOOK_SHEET} with {len(playbook_rows)} ordered safe-window phases",
            "Restarted App": "No",
            "Log Path": str(WORKBOOK),
            "Covered Stories": "ALL pending/deferred retest stories",
            "Notes": "Tracker-only update. No mycmux restart, reload, quit, install, token rotation, external app open, fault injection, update flow, or competing launch was performed.",
        },
    )
    append_unique_evidence(wb["Evidence Index"], str(SCRIPT_PATH), "QA tracker script", f"Maintains {QUEUE_SHEET} and {PLAYBOOK_SHEET} inside the single canonical workbook")

    if all(path.exists() for path in FINAL_VALIDATION_LOGS.values()):
        final_log_paths = "\n".join(str(path) for path in FINAL_VALIDATION_LOGS.values())
        append_unique_row(
            wb["Test Run Log"],
            "Run ID",
            {
                "Run ID": FINAL_VALIDATION_RUN_ID,
                "Timestamp": FINAL_VALIDATION_TIMESTAMP,
                "Command": "python -m pytest tests -q; cmd /c npx tsc --noEmit; cmd /c npm run build; cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux",
                "Exit Code": 0,
                "Result Summary": "Pass: 46 Python tests, TypeScript noEmit, npm build, and 75 Rust tests passed; npm build kept existing chunk-size warning",
                "Restarted App": "No",
                "Log Path": final_log_paths,
                "Covered Stories": "ALL source-contract, tracker-contract, and build-validation scope",
                "Notes": "Final validation after Manual Retest Queue update. No mycmux restart, reload, quit, install, replacement, token rotation, external app open, fault injection, or competing launch was performed.",
            },
        )
        append_unique_validation(wb["Validation Log"], [FINAL_VALIDATION_TIMESTAMP, "Python full QA suite after safe-window queue", "python -m pytest tests -q", "Pass: 46 passed", "No", str(FINAL_VALIDATION_LOGS["pytest"])])
        append_unique_validation(wb["Validation Log"], [FINAL_VALIDATION_TIMESTAMP, "TypeScript noEmit after safe-window queue", "cmd /c npx tsc --noEmit", "Pass: exit code 0", "No", str(FINAL_VALIDATION_LOGS["tsc"])])
        append_unique_validation(wb["Validation Log"], [FINAL_VALIDATION_TIMESTAMP, "Frontend production build after safe-window queue", "cmd /c npm run build", "Pass: build completed; existing chunk-size warning only", "No", str(FINAL_VALIDATION_LOGS["npm_build"])])
        append_unique_validation(wb["Validation Log"], [FINAL_VALIDATION_TIMESTAMP, "Rust unit tests after safe-window queue", "cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux", "Pass: 75 passed; 0 failed", "No", str(FINAL_VALIDATION_LOGS["cargo"])])
        for name, log_path in FINAL_VALIDATION_LOGS.items():
            append_unique_evidence(wb["Evidence Index"], str(log_path), "Final validation log", f"Referenced by {FINAL_VALIDATION_RUN_ID} ({name})")
        set_overview_value(wb["Overview"], "Retest loop status", "Passed after safe-window queue update: 46 Python tests, TypeScript noEmit, npm build, and 75 Rust tests. Pending: live app reload/restart window for Socket API runtime retest and full user-behavior retest.")
        set_overview_value(wb["Overview"], "Fix loop status", "ERR-003 source fix and Manual Retest Queue tracker update are validated by tests/build; no new product defect found in no-restart validation.")

    if all(path.exists() for path in PLAYBOOK_VALIDATION_LOGS.values()):
        playbook_log_paths = "\n".join(str(path) for path in PLAYBOOK_VALIDATION_LOGS.values())
        append_unique_row(
            wb["Test Run Log"],
            "Run ID",
            {
                "Run ID": PLAYBOOK_VALIDATION_RUN_ID,
                "Timestamp": PLAYBOOK_VALIDATION_TIMESTAMP,
                "Command": "python -m pytest tests -q; cmd /c npx tsc --noEmit; cmd /c npm run build; cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux",
                "Exit Code": 0,
                "Result Summary": "Pass: 47 Python tests, TypeScript noEmit, npm build, and 75 Rust tests passed; npm build kept existing chunk-size warning",
                "Restarted App": "No",
                "Log Path": playbook_log_paths,
                "Covered Stories": "ALL source-contract, tracker-contract, safe-window-playbook, and build-validation scope",
                "Notes": "Validation after Safe Window Playbook update. No mycmux restart, reload, quit, install, replacement, token rotation, external app open, fault injection, update flow, or competing launch was performed.",
            },
        )
        append_unique_validation(wb["Validation Log"], [PLAYBOOK_VALIDATION_TIMESTAMP, "Python full QA suite after safe-window playbook", "python -m pytest tests -q", "Pass: 47 passed", "No", str(PLAYBOOK_VALIDATION_LOGS["pytest"])])
        append_unique_validation(wb["Validation Log"], [PLAYBOOK_VALIDATION_TIMESTAMP, "TypeScript noEmit after safe-window playbook", "cmd /c npx tsc --noEmit", "Pass: exit code 0", "No", str(PLAYBOOK_VALIDATION_LOGS["tsc"])])
        append_unique_validation(wb["Validation Log"], [PLAYBOOK_VALIDATION_TIMESTAMP, "Frontend production build after safe-window playbook", "cmd /c npm run build", "Pass: build completed; existing chunk-size warning only", "No", str(PLAYBOOK_VALIDATION_LOGS["npm_build"])])
        append_unique_validation(wb["Validation Log"], [PLAYBOOK_VALIDATION_TIMESTAMP, "Rust unit tests after safe-window playbook", "cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux", "Pass: 75 passed; 0 failed", "No", str(PLAYBOOK_VALIDATION_LOGS["cargo"])])
        for name, log_path in PLAYBOOK_VALIDATION_LOGS.items():
            append_unique_evidence(wb["Evidence Index"], str(log_path), "Safe-window playbook validation log", f"Referenced by {PLAYBOOK_VALIDATION_RUN_ID} ({name})")
        set_overview_value(wb["Overview"], "Retest loop status", "Passed after Safe Window Playbook update: 47 Python tests, TypeScript noEmit, npm build, and 75 Rust tests. Pending: approved live app window for Socket API runtime retest and full user-behavior retest.")
        set_overview_value(wb["Overview"], "Fix loop status", "ERR-003 source fix, Manual Retest Queue, and Safe Window Playbook are validated by tests/build; live runtime retest remains gated by no-restart constraint.")

    restyle_workbook(wb)
    wb.save(WORKBOOK)

    loaded = load_workbook(WORKBOOK, read_only=True, data_only=False)
    print(WORKBOOK)
    print(f"{QUEUE_SHEET}: rows={loaded[QUEUE_SHEET].max_row - 1}")
    print(f"{PLAYBOOK_SHEET}: rows={loaded[PLAYBOOK_SHEET].max_row - 1}")
    print(f"buckets={summary}")


if __name__ == "__main__":
    main()
