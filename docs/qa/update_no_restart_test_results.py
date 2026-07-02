from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


REPO = Path(r"C:\Users\miyaz\cmux-for-linux-dev-master")
WORKBOOK = REPO / "docs" / "qa" / "mycmux-feature-status-canonical.xlsx"
LOG_DIR = REPO / "docs" / "qa" / "logs"

TIMESTAMP = "2026-06-25 00:20:00"

RUNS = [
    {
        "Run ID": "RUN-20260625-001719-PYTEST",
        "Timestamp": TIMESTAMP,
        "Command": "python -m pytest tests -q",
        "Exit Code": 0,
        "Result Summary": "Pass: 35 passed in 0.13s",
        "Restarted App": "No",
        "Log Path": str(LOG_DIR / "validation-20260625-001719-pytest.log"),
        "Covered Stories": "MYC-001, MYC-002, MYC-007, MYC-008, MYC-009, MYC-011, MYC-012, MYC-013, MYC-014, MYC-015, MYC-016, MYC-018, MYC-021, MYC-022, MYC-023, MYC-030, MYC-052",
        "Notes": "Static/contract tests only; no live app restart or competing app launch.",
    },
    {
        "Run ID": "RUN-20260625-001719-TSC",
        "Timestamp": TIMESTAMP,
        "Command": "npx tsc --noEmit",
        "Exit Code": 0,
        "Result Summary": "Pass: TypeScript typecheck completed with exit code 0",
        "Restarted App": "No",
        "Log Path": str(LOG_DIR / "validation-20260625-001719-tsc.log"),
        "Covered Stories": "MYC-004, MYC-005, MYC-006, MYC-008, MYC-009, MYC-010, MYC-011, MYC-012, MYC-013, MYC-017, MYC-020, MYC-022, MYC-024, MYC-025, MYC-026, MYC-029, MYC-031, MYC-033, MYC-034, MYC-035, MYC-036, MYC-037, MYC-038, MYC-039, MYC-040, MYC-041, MYC-042, MYC-043, MYC-044, MYC-046, MYC-047, MYC-048, MYC-049, MYC-050, MYC-051, MYC-052",
        "Notes": "Compile/type coverage only; manual UI behavior still needs runtime checks.",
    },
    {
        "Run ID": "RUN-20260625-001929-CARGO-ALL",
        "Timestamp": TIMESTAMP,
        "Command": "cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux",
        "Exit Code": 0,
        "Result Summary": "Pass: 75 Rust tests passed; 0 failed",
        "Restarted App": "No",
        "Log Path": str(LOG_DIR / "validation-20260625-001929-cargo-all.log"),
        "Covered Stories": "MYC-007, MYC-014, MYC-019, MYC-021, MYC-023, MYC-027, MYC-030, MYC-032, MYC-043, MYC-044",
        "Notes": "Rust unit tests cover storage, PTY parsing, artifact preview/edit safety, fs reveal, remote token auth, and session mapping contracts.",
    },
    {
        "Run ID": "RUN-20260625-001719-CARGO-FILTER-SUPERSEDED",
        "Timestamp": TIMESTAMP,
        "Command": "cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux usage::",
        "Exit Code": 0,
        "Result Summary": "Superseded: command matched 0 tests and was not used as proof",
        "Restarted App": "No",
        "Log Path": str(LOG_DIR / "validation-20260625-001719-cargo-usage.log"),
        "Covered Stories": "None",
        "Notes": "Logged as a QA logistics error and corrected by RUN-20260625-001929-CARGO-ALL.",
    },
    {
        "Run ID": "RUN-20260625-002449-TARGETED-STATIC",
        "Timestamp": "2026-06-25 00:24:49",
        "Command": "targeted PowerShell static checks for MYC-028, MYC-045, MYC-053..MYC-056",
        "Exit Code": 0,
        "Result Summary": "Pass: targeted static checks failCount=0",
        "Restarted App": "No",
        "Log Path": str(LOG_DIR / "validation-20260625-002449-targeted-static-v2.log"),
        "Covered Stories": "MYC-028, MYC-045, MYC-053, MYC-054, MYC-055, MYC-056",
        "Notes": "Static evidence only; user-behavior and network/runtime tests still pending where applicable.",
    },
]

COVERAGE = {
    "MYC-001": ("Automated contract pass; runtime deferred", "pytest cwd round-trip and launch cwd contract passed; restart-dependent live startup remains deferred."),
    "MYC-002": ("Automated contract pass; runtime deferred", "pytest startup restore, autosave hold, inactive workspace mount, and update-loop contract checks passed; live restart remains deferred."),
    "MYC-004": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for TitleBar/UI surface; click behavior remains manual."),
    "MYC-005": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for setup components/store usage; empty-data runtime flow remains manual."),
    "MYC-006": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for workspace tab UI; live switching remains manual."),
    "MYC-007": ("Automated contract pass; manual UI pending", "pytest and Rust storage tests passed for PersistentData/data.json storage contracts."),
    "MYC-008": ("Automated contract pass; manual UI pending", "pytest layout stability contracts passed for split structure, column reconciliation, and focus after split."),
    "MYC-009": ("Automated contract pass; manual UI pending", "pytest active-pane/surviving-tab close contracts passed; destructive UI close behavior remains manual."),
    "MYC-010": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for drag-related store/UI code; drag-and-drop runtime remains manual."),
    "MYC-011": ("Automated contract pass; manual UI pending", "pytest shortcut/focus contract around new split/tab path passed; live add-tab behavior remains manual."),
    "MYC-012": ("Automated contract pass; manual UI pending", "pytest active pane/tab follow-up contracts passed; live tab switching remains manual."),
    "MYC-013": ("Automated contract pass; manual UI pending", "pytest double-click rename focus contract passed; live rename input remains manual."),
    "MYC-014": ("Automated contract pass; manual UI pending", "pytest create_session cwd and duplicate-session idempotency contracts passed; live PTY spawn remains manual."),
    "MYC-015": ("Automated contract pass; manual UI pending", "pytest terminal input/keybinding/backpressure contracts and TypeScript noEmit passed; live I/O remains manual."),
    "MYC-016": ("Automated contract pass; manual UI pending", "pytest terminal grid, writable-size, and deferred-batch contracts passed; live resize remains manual."),
    "MYC-017": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for search-bar UI code; live search remains manual."),
    "MYC-018": ("Automated contract pass; manual UI pending", "pytest selection/copy, wheel focus, mouse sequence, and renderer stability contracts passed; live selection remains manual."),
    "MYC-019": ("Automated contract pass; manual UI pending", "Rust tests passed for shell detection, OSC7 cwd parsing, and path normalization."),
    "MYC-020": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for agent selector/store code; live agent selection remains manual."),
    "MYC-021": ("Automated contract pass; manual UI pending", "pytest and Rust session mapping tests passed for agent kind/session identity contracts."),
    "MYC-022": ("Automated contract pass; manual UI pending", "pytest Ctrl+P route/history-palette wiring contracts and TypeScript noEmit passed; live palette remains manual."),
    "MYC-023": ("Automated contract pass; manual UI pending", "pytest and Rust resume/handoff environment and sanitize contracts passed; live handoff generation remains manual."),
    "MYC-024": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for pinned roots UI/store/IPC references; live root pinning remains manual."),
    "MYC-025": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for file tree UI/store references; live directory browsing remains manual."),
    "MYC-026": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for file selection store/UI references; range selection remains manual."),
    "MYC-027": ("Automated contract pass; side-effect UI pending", "Rust fs tests passed for Explorer reveal path/select handling; launching OS Explorer remains manual/deferred."),
    "MYC-028": ("Static command pass; side-effect UI pending", "Targeted static checks passed for create_file/create_folder IPC wiring, exclusive file creation, create_dir, and leaf-name validation; disposable runtime creation remains pending."),
    "MYC-029": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for path jump/search UI references; live search remains manual."),
    "MYC-030": ("Automated contract pass; manual UI pending", "Rust artifact preview tests and pytest browser-preview persistence contract passed; live preview pane remains manual."),
    "MYC-031": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for browser pane command surface; live browser navigation remains manual."),
    "MYC-032": ("Automated contract pass; manual UI pending", "Rust editable-artifact read/save/backup/format tests passed; live editor toolbar remains manual."),
    "MYC-033": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for theme switching code; live visual switch remains manual."),
    "MYC-034": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for font settings code; live rendering remains manual."),
    "MYC-035": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for background settings code; live visual rendering remains manual."),
    "MYC-036": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for color tweak/reset code; live visual rendering remains manual."),
    "MYC-037": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for notification settings code; live toggle behavior remains manual."),
    "MYC-038": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for renderer setting code; live renderer switch remains manual."),
    "MYC-039": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for keybinding settings code; live edit behavior remains manual."),
    "MYC-040": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for notification/badge code; live badge event behavior remains manual."),
    "MYC-041": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for notification panel code; live panel behavior remains manual."),
    "MYC-042": ("Typecheck pass; manual/Rust usage pending", "TypeScript noEmit passed for usage meter UI/store code; usage aggregation-specific runtime coverage remains pending."),
    "MYC-043": ("Automated contract pass; side-effect UI pending", "Rust remote auth tests and TypeScript noEmit passed for remote access UI; live network/client display remains manual."),
    "MYC-044": ("Automated contract pass; side-effect UI pending", "Rust remote token validation/generation tests and TypeScript noEmit passed; live token rotation remains manual."),
    "MYC-045": ("Static client pass; network runtime pending", "Targeted static checks passed for remote client assets, WebSocket path, token handling, terminal app, and fit addon; live phone/browser connection remains pending."),
    "MYC-046": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for Buddy display/collapse code; live UI remains manual."),
    "MYC-047": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for Buddy settings code; live save/load remains manual."),
    "MYC-048": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for Codex Pet skin code; live skin loading remains manual."),
    "MYC-049": ("Typecheck pass; manual UI pending", "TypeScript noEmit passed for Buddy chat/memory code path; live chat remains manual."),
    "MYC-050": ("Typecheck pass; API runtime pending", "TypeScript noEmit passed for socket API types/UI references; live socket API exercise remains pending."),
    "MYC-051": ("Typecheck pass; manual fault pending", "TypeScript noEmit passed for React error-boundary code; live fault injection remains manual."),
    "MYC-052": ("Automated contract pass; update runtime deferred", "pytest confirms forced auto-update loop is not started; TypeScript noEmit passed; live update check is deferred to avoid restart/external side effects."),
    "MYC-053": ("Backlog confirmed; no runtime test", "Pending accessibility-and-motion doc exists and remains tracked as not-proven implemented; no product runtime pass claimed."),
    "MYC-054": ("Backlog confirmed; no runtime test", "Pending input-and-interaction doc exists and remains tracked as not-proven implemented; no product runtime pass claimed."),
    "MYC-055": ("Backlog confirmed; no runtime test", "Pending ssh-remote-sessions doc exists and remains tracked as not-proven implemented; no product runtime pass claimed."),
    "MYC-056": ("Backlog confirmed; no runtime test", "Pending ui-polish doc exists and remains tracked as partially unknown; no product runtime pass claimed."),
}

ERROR_ROWS = [
    [
        "ERR-001",
        "ALL",
        "Logistic",
        "Low",
        "Initial cargo filter command `cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux usage::` exited 0 but matched 0 tests, so it was weak evidence.",
        "Validation commands should prove nonzero test coverage before being used as story evidence.",
        str(LOG_DIR / "validation-20260625-001719-cargo-usage.log"),
        "Fixed",
        "List tests, then rerun the complete Rust test suite without the incorrect filter.",
        "Passed",
        "Corrected by RUN-20260625-001929-CARGO-ALL: 75 Rust tests passed.",
    ],
    [
        "ERR-002",
        "MYC-056",
        "Logistic",
        "Low",
        "First targeted static check looked for docs\\features\\pending\\ui-polish-and-advanced-surface.md, but the canonical evidence file is docs\\features\\pending\\ui-polish.md.",
        "QA verification should use the exact evidence path recorded in Feature Status.",
        str(LOG_DIR / "validation-20260625-002418-targeted-static.log"),
        "Fixed",
        "Rerun targeted static checks against the exact MYC-056 evidence path.",
        "Passed",
        "Corrected by RUN-20260625-002449-TARGETED-STATIC with failCount=0.",
    ],
]


def headers(ws):
    return [cell.value for cell in ws[1]]


def row_by_key(ws, key_name: str):
    h = headers(ws)
    key_idx = h.index(key_name) + 1
    return {ws.cell(row=row, column=key_idx).value: row for row in range(2, ws.max_row + 1)}


def append_unique(ws, key_name: str, values: dict[str, object]) -> None:
    h = headers(ws)
    existing = row_by_key(ws, key_name)
    key = values[key_name]
    if key in existing:
        row = existing[key]
    else:
        row = ws.max_row + 1
    for col, name in enumerate(h, start=1):
        ws.cell(row=row, column=col).value = values.get(name)


def append_error_unique(ws, row_values: list[object]) -> None:
    h = headers(ws)
    existing = row_by_key(ws, "Error ID")
    row = existing.get(row_values[0], ws.max_row + 1)
    for col, value in enumerate(row_values, start=1):
        ws.cell(row=row, column=col).value = value


def ensure_test_run_log(wb):
    if "Test Run Log" in wb.sheetnames:
        ws = wb["Test Run Log"]
    else:
        ws = wb.create_sheet("Test Run Log")
        ws.append(["Run ID", "Timestamp", "Command", "Exit Code", "Result Summary", "Restarted App", "Log Path", "Covered Stories", "Notes"])
    for run in RUNS:
        append_unique(ws, "Run ID", run)


def update_feature_and_matrix(wb):
    wf = wb["Feature Status"]
    wt = wb["Test Matrix"]
    feature_headers = headers(wf)
    matrix_headers = headers(wt)
    feature_rows = row_by_key(wf, "Story ID")
    matrix_rows = row_by_key(wt, "Story ID")
    evidence = "\n".join([
        str(LOG_DIR / "validation-20260625-001719-pytest.log"),
        str(LOG_DIR / "validation-20260625-001719-tsc.log"),
        str(LOG_DIR / "validation-20260625-001929-cargo-all.log"),
    ])
    for story_id, (status, note) in COVERAGE.items():
        if story_id in feature_rows:
            row = feature_rows[story_id]
            wf.cell(row=row, column=feature_headers.index("Test Status") + 1).value = status
            wf.cell(row=row, column=feature_headers.index("Error Status") + 1).value = "No product error found in no-restart automated checks"
            wf.cell(row=row, column=feature_headers.index("Fix Status") + 1).value = "No fix needed for automated checks"
            wf.cell(row=row, column=feature_headers.index("Retest Status") + 1).value = "Pending full user-behavior retest"
            existing_notes = wf.cell(row=row, column=feature_headers.index("Notes") + 1).value or ""
            wf.cell(row=row, column=feature_headers.index("Notes") + 1).value = (existing_notes + "\n" if existing_notes else "") + f"2026-06-25 no-restart automated update: {note}"
        if story_id in matrix_rows:
            row = matrix_rows[story_id]
            wt.cell(row=row, column=matrix_headers.index("Current Result") + 1).value = note
            wt.cell(row=row, column=matrix_headers.index("Evidence") + 1).value = (wt.cell(row=row, column=matrix_headers.index("Evidence") + 1).value or "") + "\n" + evidence
            wt.cell(row=row, column=matrix_headers.index("Status") + 1).value = status
            existing_notes = wt.cell(row=row, column=matrix_headers.index("Notes") + 1).value or ""
            wt.cell(row=row, column=matrix_headers.index("Notes") + 1).value = (existing_notes + "\n" if existing_notes else "") + "Automated only; manual user-behavior test still required before final completion."


def update_validation_log(wb):
    ws = wb["Validation Log"]
    rows = [
        [TIMESTAMP, "Python contract tests", "python -m pytest tests -q", "Pass: 35 passed in 0.13s", "No", str(LOG_DIR / "validation-20260625-001719-pytest.log")],
        [TIMESTAMP, "TypeScript noEmit", "npx tsc --noEmit", "Pass: exit code 0", "No", str(LOG_DIR / "validation-20260625-001719-tsc.log")],
        [TIMESTAMP, "Rust unit tests", "cargo test --manifest-path src-tauri\\Cargo.toml -p mycmux", "Pass: 75 passed; 0 failed", "No", str(LOG_DIR / "validation-20260625-001929-cargo-all.log")],
        ["2026-06-25 00:24:49", "Targeted static checks", "PowerShell static checks for MYC-028, MYC-045, MYC-053..MYC-056", "Pass: failCount=0", "No", str(LOG_DIR / "validation-20260625-002449-targeted-static-v2.log")],
        [TIMESTAMP, "No-restart constraint", "Process observation and command selection", "Pass: no mycmux restart, quit, replacement, or competing app launch was performed", "No", "Only static/type/unit/contract checks were run."],
    ]
    existing = {(ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value) for r in range(2, ws.max_row + 1)}
    for values in rows:
        key = (values[1], values[2])
        if key not in existing:
            ws.append(values)


def update_overview(wb):
    ws = wb["Overview"]
    mapping = {ws.cell(row=row, column=1).value: row for row in range(2, ws.max_row + 1)}
    updates = {
        "Current phase": "No-restart automated test pass recorded; manual UI/runtime loop still active.",
        "Runtime test loop status": "In progress: 55 stories have static/type/contract/backlog evidence; MYC-003 single-instance runtime is still not run because it risks a competing launch.",
        "Fix loop status": "Only QA logistics error ERR-001 fixed so far; no product UX/logistics defect has been proven by automated checks.",
        "Retest loop status": "Pending full user-behavior retest after manual/runtime testing and any product fixes.",
    }
    for key, value in updates.items():
        row = mapping.get(key)
        if row:
            ws.cell(row=row, column=2).value = value
        else:
            ws.append([key, value])


def style_workbook(wb):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        widths = {1: 18, 2: 22, 3: 36, 4: 54, 5: 64, 6: 24, 7: 90, 8: 54, 9: 42, 10: 28, 11: 42, 12: 24, 13: 50}
        for col_index, width in widths.items():
            if ws.max_column >= col_index:
                ws.column_dimensions[get_column_letter(col_index)].width = width
        for table in ws.tables.values():
            table.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        if ws.max_row >= 2 and not ws.tables:
            table_name = ''.join(ch for ch in ws.title if ch.isalnum())[:20] + "Table"
            table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)


def main() -> None:
    wb = load_workbook(WORKBOOK)
    ensure_test_run_log(wb)
    update_feature_and_matrix(wb)
    update_validation_log(wb)
    update_overview(wb)
    for row in ERROR_ROWS:
        append_error_unique(wb["Error Log"], row)
    style_workbook(wb)
    wb.save(WORKBOOK)
    loaded = load_workbook(WORKBOOK, read_only=True, data_only=True)
    print(WORKBOOK)
    for sheet in ["Feature Status", "Test Matrix", "Error Log", "Validation Log", "Test Run Log"]:
        print(f"{sheet}: rows={loaded[sheet].max_row - 1}")


if __name__ == "__main__":
    main()
