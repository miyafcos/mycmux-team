from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

REPO = Path(r"C:\Users\miyaz\cmux-for-linux-dev-master")
WORKBOOK = REPO / "docs" / "qa" / "mycmux-feature-status-canonical.xlsx"
LOG_DIR = REPO / "docs" / "qa" / "logs"
TIMESTAMP = "2026-06-25 01:21:00"

RUNTIME_LOG = LOG_DIR / "validation-20260625-011500-no-restart-runtime.log"
SOCKET_TIMEOUT_LOG = LOG_DIR / "validation-20260625-011650-socket-timeout-prefix.log"
PYTEST_LOG = LOG_DIR / "validation-20260625-012100-pytest-after-socket-fix.log"
TSC_LOG = LOG_DIR / "validation-20260625-012100-tsc-after-socket-fix.log"

RUNS = [
    {
        "Run ID": "RUN-20260625-011500-NORESTART-RUNTIME",
        "Timestamp": "2026-06-25 01:15:00",
        "Command": "python docs\\qa\\no_restart_runtime_checks.py",
        "Exit Code": 0,
        "Result Summary": "Pass: no-restart runtime/source checks failures=0; no secret values printed",
        "Restarted App": "No",
        "Log Path": str(RUNTIME_LOG),
        "Covered Stories": "MYC-004, MYC-005, MYC-007, MYC-024, MYC-033, MYC-034, MYC-035, MYC-036, MYC-043, MYC-045, MYC-050",
        "Notes": "Read-only process/AppData/port/HTTP checks plus source checks. No app quit, restart, install, or competing launch.",
    },
    {
        "Run ID": "RUN-20260625-011650-SOCKET-TIMEOUT-PREFIX",
        "Timestamp": "2026-06-25 01:16:50",
        "Command": "send {\"cmd\":\"list_workspaces\",\"args\":{}} to port from C:\\Users\\miyaz\\.mycmux\\mycmux.port",
        "Exit Code": 1,
        "Result Summary": "Fail reproduced: TimeoutError after 3.0s waiting for a socket response from the currently loaded app",
        "Restarted App": "No",
        "Log Path": str(SOCKET_TIMEOUT_LOG),
        "Covered Stories": "MYC-050",
        "Notes": "Product defect in loaded frontend bridge; fixed in source but live retest requires reload/restart, which is deferred by instruction.",
    },
    {
        "Run ID": "RUN-20260625-012100-PYTEST-AFTER-SOCKET-FIX",
        "Timestamp": TIMESTAMP,
        "Command": "python -m pytest tests -q",
        "Exit Code": 0,
        "Result Summary": "Pass: 36 passed in 0.09s",
        "Restarted App": "No",
        "Log Path": str(PYTEST_LOG),
        "Covered Stories": "MYC-050",
        "Notes": "Added socket API frontend bridge source contract test.",
    },
    {
        "Run ID": "RUN-20260625-012100-TSC-AFTER-SOCKET-FIX",
        "Timestamp": TIMESTAMP,
        "Command": "npx tsc --noEmit",
        "Exit Code": 0,
        "Result Summary": "Pass: TypeScript typecheck completed with exit code 0",
        "Restarted App": "No",
        "Log Path": str(TSC_LOG),
        "Covered Stories": "MYC-050",
        "Notes": "Validates new SocketListener bridge compiles.",
    },
]

UPDATES = {
    "MYC-004": {
        "status": "No-restart source check pass; live UI pending",
        "current": "TitleBar source exposes non-close controls for New Workspace, Toggle Sidebar, Notifications, and Settings. Close/minimize live click behavior remains deferred.",
        "note": "2026-06-25 no-restart manual/source pass: TitleBar controls verified without clicking close/restart paths.",
    },
    "MYC-005": {
        "status": "No-restart source check pass; live UI pending",
        "current": "WorkspaceSetup source exposes default name, grid picker, agent slot list, and onLaunch path. Empty-state live flow remains deferred.",
        "note": "2026-06-25 no-restart manual/source pass: setup wizard controls verified in source.",
    },
    "MYC-007": {
        "status": "No-restart runtime read pass; manual UI pending",
        "current": "Read-only AppData inspection parsed mycmux and mycmux-lite data.json; required keys, workspace/pane/tab shape, settings object, and pinned_roots list passed.",
        "note": "2026-06-25 no-restart runtime pass: persistent data schema validated without printing raw data or mutating files.",
    },
    "MYC-024": {
        "status": "No-restart persistence read pass; manual UI pending",
        "current": "Read-only AppData inspection confirmed pinned_roots is present as a list in both mycmux and mycmux-lite data.json; UI pin/unpin remains manual.",
        "note": "2026-06-25 no-restart runtime pass: pinned root persistence shape verified read-only.",
    },
    "MYC-033": {
        "status": "No-restart settings/source pass; manual UI pending",
        "current": "Read-only settings inspection found theme_id and theme_tweaks keys; source check verified theme store persistable fields.",
        "note": "2026-06-25 no-restart pass: theme persistence surface verified without changing active theme.",
    },
    "MYC-034": {
        "status": "No-restart settings/source pass; manual UI pending",
        "current": "Read-only settings inspection found font_family and font_size keys; source check verified theme store font fields.",
        "note": "2026-06-25 no-restart pass: font setting persistence surface verified without changing settings.",
    },
    "MYC-035": {
        "status": "No-restart source check pass; manual UI pending",
        "current": "Theme store source exposes background-related persistence surface; image path selection/live rendering remains manual.",
        "note": "2026-06-25 no-restart pass: background setting source surface verified.",
    },
    "MYC-036": {
        "status": "No-restart settings/source pass; manual UI pending",
        "current": "Read-only settings inspection found theme_tweaks; source check verified tweak persistence fields. Live color editing/reset remains manual.",
        "note": "2026-06-25 no-restart pass: color tweak persistence surface verified without changing settings.",
    },
    "MYC-043": {
        "status": "No-restart remote read pass; token rotation pending",
        "current": "Existing remote server port file is present; source check verifies Settings UI displays token_suffix and connected_clients without logging full token.",
        "note": "2026-06-25 no-restart pass: remote info display surface checked; rotate token side-effect remains deferred.",
    },
    "MYC-045": {
        "status": "No-restart remote client GET pass; WebSocket pending",
        "current": "HTTP GET / on existing remote port returned 200 text/html; /api/state without token returned 401 Unauthorized. QR/token and WebSocket terminal behavior remain deferred.",
        "note": "2026-06-25 no-restart pass: remote static client and unauthorized state guard verified without fetching QR/token.",
    },
    "MYC-050": {
        "status": "Defect reproduced and fixed in source; live retest deferred",
        "current": "External socket request list_workspaces timed out against currently loaded app. Source fix adds socket-request listener, workspace.list/list_workspaces, pane.list/list_panes, select/rename handlers, and guaranteed JSON error responses. pytest and tsc pass; live retest requires app reload/restart.",
        "note": "2026-06-25 defect ERR-003: Socket API no-response reproduced, source fixed, code retested; live runtime retest deferred by no-restart constraint.",
        "error_status": "ERR-003 documented",
        "fix_status": "Fixed in source; pytest/tsc pass",
        "retest_status": "Code retest pass; live runtime retest deferred",
    },
}

ERROR_ROW = [
    "ERR-003",
    "MYC-050",
    "Product / API runtime",
    "High",
    "Sending a documented external socket request to the running app timed out after 3 seconds instead of returning JSON.",
    "Every socket request should receive either a JSON result or a JSON error line; clients must not hang indefinitely.",
    str(SOCKET_TIMEOUT_LOG) + "\n" + str(PYTEST_LOG) + "\n" + str(TSC_LOG),
    "Fixed in source; live retest deferred",
    "Add frontend socket-request listener in SocketListener.tsx, dispatch safe workspace/pane commands, and always call socket_response on success or error.",
    "Code retest pass; live runtime pending no-restart window",
    "Current process has not loaded the patched frontend because restarting/reloading mycmux is prohibited in this session.",
]

VALIDATION_ROWS = [
    ["2026-06-25 01:15:00", "No-restart runtime/source checks", "python docs\\qa\\no_restart_runtime_checks.py", "Pass: failures=0", "No", str(RUNTIME_LOG)],
    ["2026-06-25 01:16:50", "Socket API live pre-fix probe", "send list_workspaces to current socket port", "Fail reproduced: TimeoutError timed out", "No", str(SOCKET_TIMEOUT_LOG)],
    [TIMESTAMP, "Python contract tests after Socket fix", "python -m pytest tests -q", "Pass: 36 passed in 0.09s", "No", str(PYTEST_LOG)],
    [TIMESTAMP, "TypeScript noEmit after Socket fix", "npx tsc --noEmit", "Pass: exit code 0", "No", str(TSC_LOG)],
]


def headers(ws):
    return [cell.value for cell in ws[1]]


def row_by_key(ws, key):
    h = headers(ws)
    col = h.index(key) + 1
    return {ws.cell(row=row, column=col).value: row for row in range(2, ws.max_row + 1)}


def set_row_values(ws, row, values):
    h = headers(ws)
    for key, value in values.items():
        if key in h:
            ws.cell(row=row, column=h.index(key) + 1).value = value


def append_note(existing, note):
    if not existing:
        return note
    if note in existing:
        return existing
    return f"{existing}\n{note}"


def append_unique(ws, key, values):
    rows = row_by_key(ws, key)
    row = rows.get(values[key], ws.max_row + 1)
    set_row_values(ws, row, values)


def append_validation_unique(ws, values):
    existing = {(ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value) for r in range(2, ws.max_row + 1)}
    key = (values[1], values[2])
    if key not in existing:
        ws.append(values)


def restyle_and_resize_tables(wb):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        if sheet.tables:
            for table in sheet.tables.values():
                table.ref = ref
        elif sheet.max_row >= 2:
            name = "".join(ch for ch in sheet.title if ch.isalnum())[:20] + "Table"
            table = Table(displayName=name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            sheet.add_table(table)
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            if col == 1:
                sheet.column_dimensions[letter].width = 16
            elif col in (2, 7, 8, 9, 10, 11, 12):
                sheet.column_dimensions[letter].width = 24
            else:
                sheet.column_dimensions[letter].width = 42


def main():
    wb = load_workbook(WORKBOOK)
    wf = wb["Feature Status"]
    wt = wb["Test Matrix"]
    we = wb["Error Log"]
    wv = wb["Validation Log"]
    wr = wb["Test Run Log"]

    feature_rows = row_by_key(wf, "Story ID")
    matrix_rows = row_by_key(wt, "Story ID")
    fh = headers(wf)
    th = headers(wt)

    evidence_bundle = "\n".join(str(p) for p in [RUNTIME_LOG, SOCKET_TIMEOUT_LOG, PYTEST_LOG, TSC_LOG])
    for story_id, update in UPDATES.items():
        if story_id in feature_rows:
            row = feature_rows[story_id]
            wf.cell(row=row, column=fh.index("Test Status") + 1).value = update["status"]
            wf.cell(row=row, column=fh.index("Error Status") + 1).value = update.get("error_status", "No product error found in no-restart runtime/source checks")
            wf.cell(row=row, column=fh.index("Fix Status") + 1).value = update.get("fix_status", "No fix needed for this no-restart check")
            wf.cell(row=row, column=fh.index("Retest Status") + 1).value = update.get("retest_status", "Pending full user-behavior retest")
            notes_col = fh.index("Notes") + 1
            wf.cell(row=row, column=notes_col).value = append_note(wf.cell(row=row, column=notes_col).value, update["note"])
        if story_id in matrix_rows:
            row = matrix_rows[story_id]
            wt.cell(row=row, column=th.index("Current Result") + 1).value = update["current"]
            wt.cell(row=row, column=th.index("Status") + 1).value = update["status"]
            evidence_col = th.index("Evidence") + 1
            wt.cell(row=row, column=evidence_col).value = append_note(wt.cell(row=row, column=evidence_col).value, evidence_bundle)
            notes_col = th.index("Notes") + 1
            wt.cell(row=row, column=notes_col).value = append_note(wt.cell(row=row, column=notes_col).value, update["note"])

    error_rows = row_by_key(we, "Error ID")
    row = error_rows.get(ERROR_ROW[0], we.max_row + 1)
    for col, value in enumerate(ERROR_ROW, start=1):
        we.cell(row=row, column=col).value = value

    for run in RUNS:
        append_unique(wr, "Run ID", run)
    for values in VALIDATION_ROWS:
        append_validation_unique(wv, values)

    overview = wb["Overview"]
    overview_rows = row_by_key(overview, "Field")
    overview_updates = {
        "Current phase": "No-restart runtime/source checks in progress; one Socket API product defect fixed in source; live retest deferred until safe reload/restart window.",
        "Runtime test loop status": "In progress: no-restart checks added for persistence, settings, remote static client, and socket listener presence. Restart/real UI interaction still pending.",
        "Fix loop status": "ERR-003 Socket API no-response fixed in source; no other product defects proven in this batch.",
        "Retest loop status": "pytest and tsc passed after ERR-003 source fix; live socket request retest requires app reload/restart and remains deferred.",
    }
    for key, value in overview_updates.items():
        row = overview_rows.get(key)
        if row:
            overview.cell(row=row, column=2).value = value

    restyle_and_resize_tables(wb)
    wb.save(WORKBOOK)
    wb2 = load_workbook(WORKBOOK, data_only=False)
    required = ["Feature Status", "Test Matrix", "Error Log", "Overview", "Evidence Index", "Validation Log", "Test Run Log"]
    missing = [name for name in required if name not in wb2.sheetnames]
    if missing:
        raise SystemExit(f"missing sheets: {missing}")
    print(WORKBOOK)
    print("feature_rows", wb2["Feature Status"].max_row - 1)
    print("error_rows", wb2["Error Log"].max_row - 1)
    print("test_run_rows", wb2["Test Run Log"].max_row - 1)


if __name__ == "__main__":
    main()
