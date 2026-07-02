from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


REPO = Path(r"C:\Users\miyaz\cmux-for-linux-dev-master")
WORKBOOK = REPO / "docs" / "qa" / "mycmux-feature-status-canonical.xlsx"
LOG_DIR = REPO / "docs" / "qa" / "logs"
TIMESTAMP = "2026-06-25 02:00:00"
FOLLOWUP_LOG = LOG_DIR / "validation-20260625-020000-followup-contracts.log"

FOLLOWUP_TESTS = [
    REPO / "tests" / "test_window_command_contract.py",
    REPO / "tests" / "test_no_restart_ui_surface_contract.py",
]

RUN = {
    "Run ID": "RUN-20260625-020000-FOLLOWUP-CONTRACTS",
    "Timestamp": TIMESTAMP,
    "Command": "python -m pytest tests\\test_window_command_contract.py tests\\test_no_restart_ui_surface_contract.py -q",
    "Exit Code": 0,
    "Result Summary": "Pass: 5 no-restart static contract tests passed",
    "Restarted App": "No",
    "Log Path": str(FOLLOWUP_LOG),
    "Covered Stories": "MYC-003, MYC-006, MYC-008, MYC-009, MYC-010, MYC-011, MYC-012, MYC-017, MYC-020, MYC-025, MYC-027, MYC-029, MYC-030, MYC-031, MYC-037, MYC-038, MYC-040, MYC-041, MYC-042, MYC-044, MYC-051",
    "Notes": "Source-contract follow-up only. No duplicate app launch, restart, quit, install, token rotation, external app open, or UI click was performed.",
}

STORY_UPDATES: dict[str, dict[str, str]] = {
    "MYC-003": {
        "status": "Automated window contract pass; duplicate-launch runtime deferred",
        "current": "Window IPC wrappers, Tauri command registration, leader election, reveal-main-window, and quit flushing contracts passed. Duplicate-launch live test remains deferred to avoid competing mycmux launch.",
        "retest": "Pending safe duplicate-launch runtime window",
        "note": "2026-06-25 follow-up: no-restart window command contract passed; live duplicate launch deferred by instruction.",
    },
    "MYC-006": {"status": "No-restart source contract pass; manual UI pending", "current": "TabBar/TabItem contracts for active workspace, reordering, notification count, close, rename, and New workspace button passed.", "note": "2026-06-25 follow-up: workspace tab source contract passed."},
    "MYC-008": {"status": "No-restart source contract pass; manual UI pending", "current": "PaneTabBar split controls and workspaceLayoutStore split/move contracts passed.", "note": "2026-06-25 follow-up: pane split source contract passed."},
    "MYC-009": {"status": "No-restart source contract pass; manual UI pending", "current": "PaneTabBar close-pane control and layout-store pane removal support are covered by follow-up source contracts.", "note": "2026-06-25 follow-up: pane close source contract passed."},
    "MYC-010": {"status": "No-restart source contract pass; manual drag UI pending", "current": "moveTabToPane, moveTabToSplit, movePaneToPane, movePaneToSplit, moveTabToNewWorkspace, and movePaneToNewWorkspace contracts passed.", "note": "2026-06-25 follow-up: drag/move store source contract passed."},
    "MYC-011": {"status": "No-restart source contract pass; manual UI pending", "current": "New terminal tab control and addTabToPane wiring passed follow-up source contracts.", "note": "2026-06-25 follow-up: new tab source contract passed."},
    "MYC-012": {"status": "No-restart source contract pass; manual UI pending", "current": "setActivePaneTab wiring from TerminalPane to workspaceLayoutStore passed follow-up source contracts.", "note": "2026-06-25 follow-up: pane tab activation source contract passed."},
    "MYC-017": {"status": "No-restart source contract pass; manual UI pending", "current": "SearchAddon creation/loading and find previous/next controls passed follow-up source contracts.", "note": "2026-06-25 follow-up: terminal search source contract passed."},
    "MYC-020": {"status": "No-restart source contract pass; manual launch pending", "current": "Built-in agent definitions and default agent lookup passed follow-up source contracts.", "note": "2026-06-25 follow-up: agent selection source contract passed."},
    "MYC-025": {"status": "No-restart source contract pass; manual UI pending", "current": "File explorer listDirectory/walkTree/search index contracts passed; live tree expansion remains manual.", "note": "2026-06-25 follow-up: file tree/search-index source contract passed."},
    "MYC-027": {"status": "No-restart source contract pass; OS side-effect pending", "current": "FileExplorerSidebar openWithDefault and revealInExplorer call sites passed; launching Explorer/default app remains deferred.", "note": "2026-06-25 follow-up: file open/reveal source contract passed without OS side effect."},
    "MYC-029": {"status": "No-restart source contract pass; manual UI pending", "current": "PathJumper active root, loaded entries, and search index contracts passed; Ctrl+P live operation remains manual.", "note": "2026-06-25 follow-up: path jumper source contract passed."},
    "MYC-030": {"status": "No-restart source contract pass; manual preview UI pending", "current": "Browser preview pane reload counter and openOrReloadHtmlPreviewPane contracts passed; live preview remains manual.", "note": "2026-06-25 follow-up: preview pane source contract passed."},
    "MYC-031": {"status": "No-restart source contract pass; manual preview UI pending", "current": "Browser preview tab source contracts passed; live in-app pane browsing remains manual.", "note": "2026-06-25 follow-up: browser preview source contract passed."},
    "MYC-037": {"status": "No-restart source contract pass; manual settings UI pending", "current": "Notification enablement and sound toggle contracts in SettingsMenu/settingsStore passed.", "note": "2026-06-25 follow-up: notification settings source contract passed."},
    "MYC-038": {"status": "No-restart source contract pass; manual renderer UI pending", "current": "Stable DOM terminal renderer setting surface passed follow-up source contract.", "note": "2026-06-25 follow-up: renderer toggle source contract passed."},
    "MYC-040": {"status": "No-restart source contract pass; manual UI pending", "current": "TabBar notification count and terminal inactive-notification source contracts passed.", "note": "2026-06-25 follow-up: inactive notification source contract passed."},
    "MYC-041": {"status": "No-restart source contract pass; manual UI pending", "current": "NotificationPanel list/empty/clear contracts passed; live notification history remains manual.", "note": "2026-06-25 follow-up: notification panel source contract passed."},
    "MYC-042": {"status": "No-restart source contract pass; Rust usage runtime pending", "current": "UsageMeter, UsagePopover, and get_usage_summary command contracts passed; live OAuth usage fetch remains deferred.", "note": "2026-06-25 follow-up: usage meter source contract passed."},
    "MYC-044": {"status": "No-restart source contract pass; token rotation pending", "current": "SettingsMenu remote info/token suffix/rotateRemoteToken contracts passed; token rotation side effect remains deferred.", "note": "2026-06-25 follow-up: remote token rotation source contract passed without rotating token."},
    "MYC-051": {"status": "No-restart source contract pass; manual fault pending", "current": "Layout/common ErrorBoundary and WorkspaceView wrapper contracts passed; live fault injection remains manual.", "note": "2026-06-25 follow-up: error boundary source contract passed."},
}

ERR_004 = {
    "Error ID": "ERR-004",
    "Story ID": "ALL",
    "Type": "Logistic / QA tracker",
    "Severity": "Low",
    "Observed Behavior": "Some workbook evidence and notes cells contained repeated identical lines from repeated update-script runs.",
    "Expected Behavior": "The canonical QA tracker should keep one copy of each evidence path or note line so status remains readable.",
    "Evidence": str(WORKBOOK),
    "Status": "Fixed",
    "Fix Plan": "Deduplicate repeated newline-delimited evidence and notes cells while preserving first-seen order.",
    "Retest Status": "Passed",
    "Notes": "Fixed by update_followup_contract_results.py during RUN-20260625-020000-FOLLOWUP-CONTRACTS.",
}


def headers(sheet) -> list[str]:
    return [str(cell.value) for cell in sheet[1]]


def row_by_key(sheet, key: str) -> dict[str, int]:
    hs = headers(sheet)
    col = hs.index(key) + 1
    return {str(sheet.cell(row=row, column=col).value): row for row in range(2, sheet.max_row + 1)}


def append_note(value: object, note: str) -> str:
    existing = str(value or "").strip()
    if not existing:
        return note
    if note in existing.splitlines():
        return existing
    return existing + "\n" + note


def dedupe_multiline(value: object) -> object:
    if not isinstance(value, str) or "\n" not in value:
        return value
    seen: set[str] = set()
    kept: list[str] = []
    for line in value.splitlines():
        key = line.strip()
        if not key:
            continue
        if key in seen:
            continue
        seen.add(key)
        kept.append(line)
    return "\n".join(kept)


def append_unique_row(sheet, key_col: str, values_by_header: dict[str, object]) -> None:
    hs = headers(sheet)
    rows = row_by_key(sheet, key_col)
    key = str(values_by_header[key_col])
    row = rows.get(key, sheet.max_row + 1)
    for col, header in enumerate(hs, start=1):
        if header in values_by_header:
            sheet.cell(row=row, column=col).value = values_by_header[header]


def append_validation(sheet, values: list[object]) -> None:
    existing = {(sheet.cell(row=row, column=2).value, sheet.cell(row=row, column=3).value) for row in range(2, sheet.max_row + 1)}
    key = (values[1], values[2])
    if key not in existing:
        sheet.append(values)


def append_evidence(sheet, source: str, kind: str, notes: str) -> None:
    existing = {str(sheet.cell(row=row, column=1).value) for row in range(2, sheet.max_row + 1)}
    if source not in existing:
        sheet.append([source, kind, notes])


def restyle_and_resize_tables(wb) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        if sheet.tables:
            for table in sheet.tables.values():
                table.ref = ref
        elif sheet.max_row >= 2:
            table_name = "".join(ch for ch in sheet.title if ch.isalnum())[:20] + "Table"
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            sheet.add_table(table)
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            if col == 1:
                sheet.column_dimensions[letter].width = 16
            elif col in (2, 7, 8, 9, 10, 11, 12):
                sheet.column_dimensions[letter].width = 24
            else:
                sheet.column_dimensions[letter].width = 42
        if sheet.title == "Overview":
            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 130


def main() -> None:
    wb = load_workbook(WORKBOOK)
    feature = wb["Feature Status"]
    matrix = wb["Test Matrix"]
    error_log = wb["Error Log"]
    validation = wb["Validation Log"]
    test_runs = wb["Test Run Log"]
    evidence = wb["Evidence Index"]

    feature_rows = row_by_key(feature, "Story ID")
    matrix_rows = row_by_key(matrix, "Story ID")
    fh = headers(feature)
    mh = headers(matrix)
    evidence_bundle = "\n".join(str(path) for path in [*FOLLOWUP_TESTS, FOLLOWUP_LOG])

    for story_id, update in STORY_UPDATES.items():
        if story_id in feature_rows:
            row = feature_rows[story_id]
            feature.cell(row=row, column=fh.index("Test Status") + 1).value = update["status"]
            feature.cell(row=row, column=fh.index("Error Status") + 1).value = "No product error found in no-restart source contract checks"
            feature.cell(row=row, column=fh.index("Fix Status") + 1).value = "No product fix needed for this source-contract check"
            feature.cell(row=row, column=fh.index("Retest Status") + 1).value = update.get("retest", "Pending full user-behavior retest")
            note_col = fh.index("Notes") + 1
            feature.cell(row=row, column=note_col).value = append_note(feature.cell(row=row, column=note_col).value, update["note"])
        if story_id in matrix_rows:
            row = matrix_rows[story_id]
            matrix.cell(row=row, column=mh.index("Current Result") + 1).value = update["current"]
            matrix.cell(row=row, column=mh.index("Status") + 1).value = update["status"]
            evidence_col = mh.index("Evidence") + 1
            matrix.cell(row=row, column=evidence_col).value = append_note(matrix.cell(row=row, column=evidence_col).value, evidence_bundle)
            note_col = mh.index("Notes") + 1
            matrix.cell(row=row, column=note_col).value = append_note(matrix.cell(row=row, column=note_col).value, update["note"])

    append_unique_row(test_runs, "Run ID", RUN)
    append_validation(validation, [TIMESTAMP, "Follow-up no-restart source contracts", RUN["Command"], RUN["Result Summary"], "No", str(FOLLOWUP_LOG)])
    append_validation(validation, [TIMESTAMP, "QA tracker duplicate cleanup", "Deduplicate newline-delimited Evidence/Notes cells", "Pass: repeated lines removed while preserving first-seen order", "No", str(WORKBOOK)])
    for path in FOLLOWUP_TESTS:
        append_evidence(evidence, str(path), "No-restart contract test", "Referenced by RUN-20260625-020000-FOLLOWUP-CONTRACTS")
    append_evidence(evidence, str(FOLLOWUP_LOG), "Validation log", "pytest output for no-restart follow-up contract tests")

    append_unique_row(error_log, "Error ID", ERR_004)

    for sheet_name, cols in {
        "Feature Status": ["Notes"],
        "Test Matrix": ["Evidence", "Notes"],
        "Error Log": ["Evidence", "Notes"],
    }.items():
        sheet = wb[sheet_name]
        hs = headers(sheet)
        for col_name in cols:
            col = hs.index(col_name) + 1
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=col).value = dedupe_multiline(sheet.cell(row=row, column=col).value)

    overview = wb["Overview"]
    overview_rows = row_by_key(overview, "Field")
    overview_updates = {
        "Current phase": "No-restart runtime/source checks continued; MYC-003 is now covered by source contract; manual and restart-dependent behavior remains open.",
        "Runtime test loop status": "In progress: follow-up source contracts added for remaining P0/P1 UI, side-effect, usage, notification, preview, and window stories. Restart/real UI interaction still pending.",
        "Fix loop status": "ERR-003 fixed in source; ERR-004 QA tracker duplicate lines fixed; no new product defect found in this follow-up.",
        "Retest loop status": "Follow-up source contracts passed; full user-behavior retest and live Socket API duplicate/restart-sensitive checks remain deferred.",
    }
    for key, value in overview_updates.items():
        row = overview_rows.get(key)
        if row:
            overview.cell(row=row, column=2).value = value

    restyle_and_resize_tables(wb)
    wb.save(WORKBOOK)

    loaded = load_workbook(WORKBOOK, read_only=True, data_only=False)
    required = ["Feature Status", "Test Matrix", "Error Log", "Overview", "Evidence Index", "Validation Log", "Test Run Log"]
    missing = [name for name in required if name not in loaded.sheetnames]
    if missing:
        raise SystemExit(f"missing sheets: {missing}")
    print(WORKBOOK)
    for sheet_name in required:
        print(f"{sheet_name}: rows={loaded[sheet_name].max_row - 1}")


if __name__ == "__main__":
    main()

